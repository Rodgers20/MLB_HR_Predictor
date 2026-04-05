"""
Prediction History page.
Shows full prediction log with actual HR results, player photos, and outcome badges.
Fixes: adds Opponent column, proper HR display (Pending / 0 HR / N HR), photo thumbnails.
"""

from pathlib import Path

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from dash import Input, Output, callback, dash_table, dcc, html
from openpyxl import load_workbook

EXCEL_PATH = Path("MLB_HR_Predictions.xlsx")

PRED_COLUMNS = [
    "Date", "Player", "Team", "Opponent", "Pitcher",
    "HR_Probability", "Confidence", "Park_Factor",
    "Temp_F", "Wind_Speed_MPH", "Wind_Direction", "Is_Indoor",
    "Home_Game", "Predicted_HRs", "Actual_HRs", "Hit",
]

CONF_COLORS = {"High": "#22c55e", "Medium": "#ff6b00", "Low": "#8e909c"}


# ─── Layout ───────────────────────────────────────────────────────────────────

def layout():
    return html.Div([
        html.Div([
            html.Div([
                html.H2("Prediction History", className="page-title"),
                html.P("Complete log of all predictions with actual HR results",
                       className="page-subtitle"),
            ]),
        ], className="page-header"),

        # ── Filters ──────────────────────────────────────────────────────────
        html.Div([
            html.Div([
                html.Label("Date Range"),
                dcc.DatePickerRange(id="hist-date-range"),
            ]),
            html.Div([
                html.Label("Confidence"),
                dcc.Dropdown(
                    id="hist-conf-filter",
                    options=[
                        {"label": "All",    "value": "All"},
                        {"label": "High",   "value": "High"},
                        {"label": "Medium", "value": "Medium"},
                        {"label": "Low",    "value": "Low"},
                    ],
                    value="All", clearable=False,
                    style={"width": "150px", "fontSize": "13px"},
                ),
            ]),
            html.Div([
                html.Label("Result"),
                dcc.Dropdown(
                    id="hist-result-filter",
                    options=[
                        {"label": "All",         "value": "All"},
                        {"label": "Hits Only",    "value": "Hit"},
                        {"label": "Misses Only",  "value": "Miss"},
                        {"label": "Pending",      "value": "Pending"},
                    ],
                    value="All", clearable=False,
                    style={"width": "150px", "fontSize": "13px"},
                ),
            ]),
            html.Div([
                html.Label("Search Player"),
                dcc.Input(id="hist-player-search", placeholder="Player name…",
                          debounce=True, className="search-input",
                          style={"width": "180px"}),
            ]),
        ], className="filters-row"),

        # ── Summary bar ──────────────────────────────────────────────────────
        html.Div(id="hist-summary-cards", className="summary-cards"),

        # ── History table ─────────────────────────────────────────────────────
        html.Div([
            html.Div([
                html.H4("Prediction Log", style={"margin": 0}),
                html.Span(id="hist-row-count",
                          style={"fontSize": "12px", "color": "#8e909c"}),
            ], className="table-title-bar"),

            dash_table.DataTable(
                id="history-table",
                columns=[
                    {"name": "Date",       "id": "Date"},
                    {"name": "Player",     "id": "Player"},
                    {"name": "Team",       "id": "Team"},
                    {"name": "Opponent",   "id": "Opponent"},
                    {"name": "Pitcher",    "id": "Pitcher"},
                    {"name": "HR Prob",    "id": "HR_Probability", "type": "numeric",
                     "format": {"specifier": ".1%"}},
                    {"name": "Confidence", "id": "Confidence"},
                    {"name": "Actual HRs", "id": "HR_Result"},
                    {"name": "Outcome",    "id": "Outcome_Display"},
                ],
                data=[],
                sort_action="native",
                filter_action="native",
                page_size=30,
                export_format="csv",
                style_table={"overflowX": "auto"},
                style_header={
                    "backgroundColor": "#1c1c1c",
                    "color": "#f2f2f2",
                    "fontWeight": "700",
                    "textAlign": "center",
                    "padding": "11px 10px",
                    "fontSize": "12px",
                    "letterSpacing": ".4px",
                    "textTransform": "uppercase",
                    "borderBottom": "2px solid #ff6b00",
                    "border": "none",
                },
                style_cell={
                    "textAlign": "center",
                    "padding": "9px 12px",
                    "fontFamily": "Manrope, Inter, sans-serif",
                    "fontSize": "13px",
                    "backgroundColor": "#151515",
                    "color": "#f2f2f2",
                    "border": "1px solid rgba(255,255,255,.06)",
                },
                style_cell_conditional=[
                    {"if": {"column_id": "Player"},         "textAlign": "left", "fontWeight": "600"},
                    {"if": {"column_id": "Date"},           "width": "100px"},
                    {"if": {"column_id": "Team"},           "width": "60px"},
                    {"if": {"column_id": "Opponent"},       "width": "80px"},
                    {"if": {"column_id": "HR_Result"},      "width": "90px", "fontWeight": "700"},
                    {"if": {"column_id": "Outcome_Display"},"width": "130px"},
                    {"if": {"column_id": "Confidence"},     "width": "90px"},
                ],
                style_data_conditional=[
                    # Confidence tiers
                    {"if": {"filter_query": '{Confidence} = "High"'},
                     "backgroundColor": "rgba(34,197,94,.08)"},
                    {"if": {"filter_query": '{Confidence} = "Medium"'},
                     "backgroundColor": "rgba(255,107,0,.08)"},
                    # Outcome highlighting
                    {"if": {"filter_query": '{Outcome_Display} contains "HR Hit"'},
                     "backgroundColor": "rgba(34,197,94,.12)", "color": "#86efac", "fontWeight": "700"},
                    {"if": {"filter_query": '{Outcome_Display} contains "HR"',
                            "column_id": "HR_Result"},
                     "fontWeight": "700", "color": "#86efac"},
                    {"if": {"filter_query": '{Outcome_Display} = "✗ No HR"'},
                     "backgroundColor": "rgba(239,68,68,.1)", "color": "#fca5a5"},
                    {"if": {"filter_query": '{Outcome_Display} = "⏳ Pending"'},
                     "color": "#93c5fd", "fontStyle": "italic"},
                    {"if": {"row_index": "odd"}, "backgroundColor": "#101416"},
                ],
            ),
        ], className="table-container"),

        # ── Monthly accuracy chart ────────────────────────────────────────────
        html.Div([
            html.H3("Monthly Hit Rate by Confidence", className="section-title"),
            dcc.Graph(id="monthly-perf-chart", config={"displayModeBar": False}),
        ], className="chart-section"),

        # ── HR Hitters showcase ───────────────────────────────────────────────
        html.Div([
            html.H3("HR Hitters — Confirmed Results", className="section-title"),
            html.Div(id="hr-showcase", className="picks-grid"),
        ], className="chart-section"),

        dcc.Interval(id="hist-refresh", interval=10 * 60 * 1000, n_intervals=0),
    ], className="page-container")


# ─── Data loader ──────────────────────────────────────────────────────────────

def _load_history() -> pd.DataFrame:
    if not EXCEL_PATH.exists():
        return pd.DataFrame(columns=PRED_COLUMNS)
    try:
        wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        ws = wb["Predictions"]
        data = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
        if not data:
            return pd.DataFrame(columns=PRED_COLUMNS)
        # Trim/pad rows to match PRED_COLUMNS length so mismatched Excel columns don't fail
        n = len(PRED_COLUMNS)
        data = [row[:n] for row in data]
        df = pd.DataFrame(data, columns=PRED_COLUMNS)
        # Only keep rows that are actual model predictions (have a valid HR_Probability)
        df = df[df["Player"].notna()]
        df["HR_Probability"] = pd.to_numeric(df["HR_Probability"], errors="coerce")
        df = df[df["HR_Probability"].notna()]
        return df
    except Exception:
        return pd.DataFrame(columns=PRED_COLUMNS)


def _is_pending(row) -> bool:
    """True when no actual result has been recorded yet."""
    actual = row.get("Actual_HRs")
    hit    = row.get("Hit")
    return (actual is None or actual == "" or str(actual).strip() == "") and \
           (hit    is None or hit    == "" or str(hit).strip()    == "")


SEASON_YEAR  = 2026
SEASON_START = f"{SEASON_YEAR}-01-01"


# ─── Callback ─────────────────────────────────────────────────────────────────

@callback(
    Output("history-table", "data"),
    Output("monthly-perf-chart", "figure"),
    Output("hist-summary-cards", "children"),
    Output("hist-row-count", "children"),
    Output("hr-showcase", "children"),
    Input("hist-refresh", "n_intervals"),
    Input("hist-date-range", "start_date"),
    Input("hist-date-range", "end_date"),
    Input("hist-conf-filter", "value"),
    Input("hist-result-filter", "value"),
    Input("hist-player-search", "value"),
)
def update_history(_n, start_date, end_date, conf_filter, result_filter, player_search):
    df = _load_history()
    if df.empty:
        return [], _empty_fig("No prediction history yet"), [], "", []

    df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.strftime("%Y-%m-%d")

    # ── Filters ───────────────────────────────────────────────────────────────
    # Default to current season when no date range is selected
    effective_start = start_date if start_date else SEASON_START
    df = df[df["Date"] >= effective_start]
    if end_date:
        df = df[df["Date"] <= end_date]
    if conf_filter != "All":
        df = df[df["Confidence"] == conf_filter]
    if player_search:
        df = df[df["Player"].str.contains(player_search, case=False, na=False)]

    # ── Outcome columns ───────────────────────────────────────────────────────
    df["_pending"] = df.apply(_is_pending, axis=1)
    df["Actual_HRs_num"] = pd.to_numeric(df["Actual_HRs"], errors="coerce")

    # HR_Result: plain numeric or "—" for pending
    df["HR_Result"] = df.apply(
        lambda r: "—" if r["_pending"] else int(r["Actual_HRs_num"])
        if pd.notna(r["Actual_HRs_num"]) else "—",
        axis=1,
    )

    # Outcome_Display: ✓ / ✗ / ⏳
    def _outcome(row):
        if row["_pending"]:
            return "⏳ Pending"
        hrs = row["Actual_HRs_num"]
        if pd.isna(hrs):
            return "⏳ Pending"
        return f"✓ HR Hit ({int(hrs)} HR)" if int(hrs) > 0 else "✗ No HR"

    df["Outcome_Display"] = df.apply(_outcome, axis=1)

    # ── Result filter (applied after building Outcome_Display) ────────────────
    if result_filter == "Hit":
        df = df[df["Actual_HRs_num"] > 0]
    elif result_filter == "Miss":
        df = df[(~df["_pending"]) & (df["Actual_HRs_num"] == 0)]
    elif result_filter == "Pending":
        df = df[df["_pending"]]

    # ── Table data ────────────────────────────────────────────────────────────
    table_cols = [
        "Date", "Player", "Team", "Opponent", "Pitcher",
        "HR_Probability", "Confidence",
        "HR_Result", "Outcome_Display",
    ]
    table_data = df[table_cols].to_dict("records")

    # ── Summary cards ─────────────────────────────────────────────────────────
    summary = _build_summary(df)

    row_count = f"{len(df):,} rows"

    # ── HR showcase (players who actually hit HRs — current season only) ────────
    SEASON_YEAR = 2026
    hr_hitters = df[df["Actual_HRs_num"] > 0].copy()
    hr_hitters["_year"] = pd.to_datetime(hr_hitters["Date"], errors="coerce").dt.year
    hr_hitters = hr_hitters[hr_hitters["_year"] == SEASON_YEAR]
    # One card per player-game: keep highest Actual_HRs if multiple prediction rows share the same date
    hr_hitters = (
        hr_hitters
        .sort_values("Actual_HRs_num", ascending=False)
        .drop_duplicates(subset=["Player", "Date"], keep="first")
        .sort_values("Actual_HRs_num", ascending=False)
    )
    showcase = _build_hr_showcase(hr_hitters.head(12))

    return table_data, _monthly_chart(df), summary, row_count, showcase


# ─── HR showcase cards ────────────────────────────────────────────────────────

def _build_hr_showcase(df: pd.DataFrame):
    if df.empty:
        return [html.Div([
            html.Div("⚾", className="empty-state-icon"),
            html.Div("No confirmed HR hitters yet", className="empty-state-msg"),
            html.Div("Results are filled in after games complete", className="empty-state-sub"),
        ], className="empty-state")]

    try:
        from utils.player_photos import batch_headshot_urls
        photo_map = batch_headshot_urls(df["Player"].tolist(), width=90)
    except Exception:
        photo_map = {}

    cards = []
    for _, row in df.iterrows():
        name   = str(row.get("Player", ""))
        team   = str(row.get("Team", ""))
        date_s = str(row.get("Date", ""))
        hrs    = int(row.get("Actual_HRs_num", 1))
        prob   = row.get("HR_Probability", 0) or 0
        photo  = photo_map.get(name, "")

        hr_label = f"{hrs} HR" + ("s" if hrs > 1 else "")

        card = html.Div([
            html.Img(src=photo, alt=name, className="pick-card-photo",
                     style={"borderColor": "#22c55e"}),
            html.P(name, className="pick-card-name"),
            html.P(team, className="pick-card-team"),
            html.Div(hr_label,
                     style={"fontFamily": "Manrope, sans-serif",
                            "fontSize": "26px", "fontWeight": "800",
                            "color": "#86efac", "margin": "4px 0",
                            "textShadow": "0 0 12px rgba(34,197,94,.4)"}),
            html.P(f"Predicted: {prob:.1%}", className="pick-card-matchup"),
            html.P(date_s, style={"fontSize": "11px", "color": "#64748b", "margin": "2px 0"}),
            html.Span("✓ Confirmed", className="badge badge-hit",
                      style={"marginTop": "6px"}),
        ], className="pick-card high")
        cards.append(card)

    return cards


# ─── Summary cards ────────────────────────────────────────────────────────────

def _build_summary(df: pd.DataFrame):
    total    = len(df)
    pending  = int(df["_pending"].sum())
    resolved = total - pending
    hits     = int((df["Actual_HRs_num"] > 0).sum()) if resolved > 0 else 0
    hit_rate = hits / resolved if resolved > 0 else 0

    # High-confidence accuracy
    high_df   = df[(df["Confidence"] == "High") & (~df["_pending"])]
    high_hits = int((high_df["Actual_HRs_num"] > 0).sum()) if not high_df.empty else 0
    high_rate = high_hits / len(high_df) if not high_df.empty else 0

    def card(title, value, cls=""):
        return html.Div([
            html.Div(str(value), className="summary-card-value",
                     style={"color": "#22c55e" if cls == "green" else
                            "#60a5fa" if cls == "blue" else
                            "#ff6b00" if cls == "amber" else
                            "#ff6b00"}),
            html.Div(title, className="summary-card-label"),
        ], className=f"summary-card {cls}")

    return [
        card("Total Predictions", total),
        card("HR Hits", hits, "green"),
        card("Overall Hit Rate", f"{hit_rate:.1%}", "amber"),
        card("High Conf. Accuracy", f"{high_rate:.1%}", "green"),
        card("Pending Results", pending, "blue"),
    ]


# ─── Monthly chart ────────────────────────────────────────────────────────────

def _monthly_chart(df: pd.DataFrame):
    resolved = df[~df["_pending"]].copy()
    if resolved.empty:
        return _empty_fig("No completed results yet")

    resolved["Month"] = pd.to_datetime(resolved["Date"], errors="coerce") \
                          .dt.to_period("M").astype(str)
    resolved["hit_num"] = (resolved["Actual_HRs_num"] > 0).astype(int)

    monthly = (
        resolved.groupby(["Month", "Confidence"])
        .agg(total=("hit_num", "count"), hits=("hit_num", "sum"))
        .reset_index()
    )
    monthly["hit_rate"] = monthly["hits"] / monthly["total"]

    fig = px.bar(
        monthly, x="Month", y="hit_rate", color="Confidence",
        barmode="group",
        color_discrete_map=CONF_COLORS,
        labels={"hit_rate": "Hit Rate", "Month": "Month"},
        text_auto=".0%",
    )
    fig.update_layout(
        paper_bgcolor="#151515", plot_bgcolor="#101416",
        yaxis_tickformat=".0%", height=300,
        margin=dict(l=60, r=20, t=20, b=40),
        legend_title_text="Confidence",
        font={"family": "Manrope, Inter, sans-serif", "size": 12, "color": "#f2f2f2"},
        xaxis={"gridcolor": "rgba(255,255,255,.06)", "color": "#8e909c"},
        yaxis={"gridcolor": "rgba(255,255,255,.06)", "color": "#8e909c"},
    )
    fig.update_traces(textposition="outside")
    return fig


def _empty_fig(msg: str):
    fig = go.Figure()
    fig.add_annotation(text=msg, xref="paper", yref="paper", x=0.5, y=0.5,
                       showarrow=False, font={"size": 14, "color": "#64748b"})
    fig.update_layout(height=300, paper_bgcolor="#151515", plot_bgcolor="#101416",
                      margin=dict(l=40, r=20, t=20, b=40))
    return fig

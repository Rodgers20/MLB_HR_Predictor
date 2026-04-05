"""
Model Performance page.
Reads directly from the Predictions sheet so charts reflect every resolved game.

Key additions vs the basic version:
  1. Red warning banner  — fires when High-confidence hit rate (last 14 days) < 15 %
  2. Confidence calibration chart  — proves (or disproves) that High > Medium > Low hit rates
  3. Rolling 7-day accuracy lines  — spot trends before they become crises
"""

from pathlib import Path

import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from dash import Input, Output, callback, dcc, html
from openpyxl import load_workbook

EXCEL_PATH = Path("MLB_HR_Predictions.xlsx")

PRED_COLUMNS = [
    "Date", "Player", "Team", "Opponent", "Pitcher",
    "HR_Probability", "Confidence", "Park_Factor",
    "Temp_F", "Wind_Speed_MPH", "Wind_Direction", "Is_Indoor",
    "Home_Game", "Predicted_HRs", "Actual_HRs", "Hit",
    "Insight_Text",
]
PERF_COLUMNS = [
    "Date", "Total_Predictions", "High_Conf_Count",
    "Daily_Accuracy", "Cumulative_Accuracy",
    "High_Conf_Hit_Rate", "ROI_Flat_Bet",
]

WARN_THRESHOLD = 0.15   # red banner if High-conf hit rate < 15 %
CONF_COLORS    = {"High": "#22c55e", "Medium": "#ff6b00", "Low": "#8e909c"}


# ─── Layout ───────────────────────────────────────────────────────────────────

def layout():
    return html.Div([
        html.Div([
            html.Div([
                html.H2("Model Performance", className="page-title"),
                html.P("Accuracy trends, calibration, and ROI tracking",
                       className="page-subtitle"),
            ]),
        ], className="page-header"),

        # Warning banner — hidden until callback fires
        html.Div(id="perf-warning-banner"),

        # KPI cards
        html.Div(id="perf-kpi-cards", className="summary-cards"),

        # ── Row 1: Rolling accuracy  +  Calibration ───────────────────────────
        html.Div([
            html.Div([
                html.Div([
                    html.H3("Rolling 7-Day Accuracy by Confidence Tier",
                            className="section-title"),
                    html.Span("Are predictions improving over time?",
                              style={"fontSize": "12px", "color": "#8e909c",
                                     "marginBottom": "10px", "display": "block"}),
                    dcc.Graph(id="rolling-accuracy-chart",
                              config={"displayModeBar": False}),
                ], className="chart-section", style={"flex": "1"}),

                html.Div([
                    html.H3("Confidence Calibration",
                            className="section-title"),
                    html.Span("High should beat Medium, Medium should beat Low",
                              style={"fontSize": "12px", "color": "#8e909c",
                                     "marginBottom": "10px", "display": "block"}),
                    dcc.Graph(id="calibration-chart",
                              config={"displayModeBar": False}),
                ], className="chart-section", style={"flex": "1"}),
            ], style={"display": "flex", "gap": "16px", "flexWrap": "wrap"}),
        ]),

        # ── Row 2: Cumulative accuracy  +  ROI ────────────────────────────────
        html.Div([
            html.Div([
                html.H3("Cumulative Accuracy Over Time", className="section-title"),
                dcc.Graph(id="accuracy-chart", config={"displayModeBar": False}),
            ], className="chart-section", style={"flex": "1"}),

            html.Div([
                html.H3("Cumulative ROI (flat $1 on High Confidence picks)",
                        className="section-title"),
                dcc.Graph(id="roi-chart", config={"displayModeBar": False}),
            ], className="chart-section", style={"flex": "1"}),
        ], style={"display": "flex", "gap": "16px", "flexWrap": "wrap"}),

        # ── Feature importance ────────────────────────────────────────────────
        html.Div([
            html.H3("Top Feature Importances (SHAP)", className="section-title"),
            html.Span("Which inputs drive the model's HR probability scores",
                      style={"fontSize": "12px", "color": "#8e909c",
                             "marginBottom": "10px", "display": "block"}),
            dcc.Graph(id="feature-importance-chart", config={"displayModeBar": False}),
        ], className="chart-section"),

        dcc.Interval(id="perf-refresh", interval=10 * 60 * 1000, n_intervals=0),
    ], className="page-container")


# ─── Data loaders ─────────────────────────────────────────────────────────────

def _load_predictions() -> pd.DataFrame:
    """Load Predictions sheet — only rows with resolved results."""
    if not EXCEL_PATH.exists():
        return pd.DataFrame(columns=PRED_COLUMNS)
    try:
        wb   = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        ws   = wb["Predictions"]
        data = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
    except Exception:
        return pd.DataFrame(columns=PRED_COLUMNS)

    if not data:
        return pd.DataFrame(columns=PRED_COLUMNS)

    row_len = len(data[0]) if data else 0
    cols = PRED_COLUMNS[:row_len] if row_len <= len(PRED_COLUMNS) else PRED_COLUMNS
    df = pd.DataFrame(data, columns=cols)
    for missing in PRED_COLUMNS:
        if missing not in df.columns:
            df[missing] = ""
    df = df[df["Player"].notna()].copy()
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df["HR_Probability"] = pd.to_numeric(df["HR_Probability"], errors="coerce")
    df["Actual_HRs_num"] = pd.to_numeric(df["Actual_HRs"], errors="coerce")
    df["Hit_num"] = (df["Actual_HRs_num"] > 0).astype(float)

    # Keep only rows with confirmed results
    df = df[df["Actual_HRs_num"].notna()].copy()

    # Filter out predictions from the old uncalibrated model (probability > 21%).
    # Those were generated before the dampening/cap fix and skew calibration charts.
    # Stub rows (N/A confidence, prob = 0) are kept because they carry actual results.
    model_rows = df[df["Confidence"] != "N/A"]
    stub_rows  = df[df["Confidence"] == "N/A"]
    model_rows = model_rows[model_rows["HR_Probability"] <= 0.21]
    df = pd.concat([model_rows, stub_rows], ignore_index=True)
    return df


def _load_performance() -> pd.DataFrame:
    if not EXCEL_PATH.exists():
        return pd.DataFrame(columns=PERF_COLUMNS)
    try:
        wb   = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        ws   = wb["Model_Performance"]
        data = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
        if not data:
            return pd.DataFrame(columns=PERF_COLUMNS)
        df = pd.DataFrame(data, columns=PERF_COLUMNS).dropna(subset=["Date"])
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
        return df
    except Exception:
        return pd.DataFrame(columns=PERF_COLUMNS)


def _load_feature_importance() -> pd.DataFrame:
    if not EXCEL_PATH.exists():
        return pd.DataFrame(columns=["Date", "Feature", "Importance_Score"])
    try:
        wb   = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        ws   = wb["Feature_Importance"]
        data = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
        if not data:
            return pd.DataFrame(columns=["Date", "Feature", "Importance_Score"])
        return pd.DataFrame(data, columns=["Date", "Feature", "Importance_Score"])
    except Exception:
        return pd.DataFrame(columns=["Date", "Feature", "Importance_Score"])


# ─── Callback ─────────────────────────────────────────────────────────────────

@callback(
    Output("perf-warning-banner",       "children"),
    Output("perf-kpi-cards",            "children"),
    Output("rolling-accuracy-chart",    "figure"),
    Output("calibration-chart",         "figure"),
    Output("accuracy-chart",            "figure"),
    Output("roi-chart",                 "figure"),
    Output("feature-importance-chart",  "figure"),
    Input("perf-refresh", "n_intervals"),
)
def update_performance(_n):
    preds   = _load_predictions()
    perf_df = _load_performance()
    fi_df   = _load_feature_importance()

    warning  = _warning_banner(preds)
    cards    = _build_kpi_cards(preds, perf_df)
    rolling  = _rolling_accuracy_chart(preds)
    calib    = _calibration_chart(preds)
    acc      = _accuracy_chart(perf_df)
    roi      = _roi_chart(perf_df)
    feat     = _feature_importance_chart(fi_df)

    return warning, cards, rolling, calib, acc, roi, feat


# ─── Warning banner ───────────────────────────────────────────────────────────

def _warning_banner(df: pd.DataFrame):
    if df.empty:
        return []

    # Evaluate only last 14 days to reflect recent performance
    cutoff   = df["Date"].max() - pd.Timedelta(days=14)
    recent   = df[(df["Date"] >= cutoff) & (df["Confidence"] == "High")]

    if recent.empty:
        return []

    hit_rate = float(recent["Hit_num"].mean())

    if hit_rate >= WARN_THRESHOLD:
        return []   # No warning needed

    n = len(recent)
    return html.Div([
        html.Div([
            html.Span("⚠", style={"fontSize": "24px", "marginRight": "12px"}),
            html.Div([
                html.Strong("Model Alert: High-Confidence Hit Rate Below Threshold"),
                html.Div(
                    f"High-confidence picks have hit at {hit_rate:.1%} over the last 14 days "
                    f"({n} picks) — below the {WARN_THRESHOLD:.0%} minimum. "
                    "Consider reviewing feature weights or retraining the model.",
                    style={"fontSize": "13px", "marginTop": "4px", "opacity": ".9"},
                ),
            ]),
        ], style={
            "display": "flex", "alignItems": "flex-start",
            "background": "#7f1d1d",
            "color": "white",
            "border": "1px solid #991b1b",
            "borderRadius": "10px",
            "padding": "16px 20px",
            "marginBottom": "20px",
            "boxShadow": "0 4px 12px rgba(239,68,68,.25)",
        }),
    ])


# ─── KPI cards ────────────────────────────────────────────────────────────────

def _build_kpi_cards(preds: pd.DataFrame, perf_df: pd.DataFrame):
    if preds.empty:
        return [html.Div(
            "No resolved predictions yet — run the scheduler and wait for game results.",
            style={"color": "#8e909c", "padding": "20px", "fontStyle": "italic"},
        )]

    total        = len(preds)
    overall_rate = float(preds["Hit_num"].mean())
    high_df      = preds[preds["Confidence"] == "High"]
    high_rate    = float(high_df["Hit_num"].mean()) if not high_df.empty else 0.0
    days         = preds["Date"].nunique()

    cum_roi = 0.0
    if not perf_df.empty and "ROI_Flat_Bet" in perf_df.columns:
        cum_roi = float(perf_df["ROI_Flat_Bet"].sum())

    # Last-7-day trend
    cutoff   = preds["Date"].max() - pd.Timedelta(days=7)
    recent   = preds[preds["Date"] >= cutoff]
    last7    = float(recent["Hit_num"].mean()) if not recent.empty else 0.0
    trend    = last7 - overall_rate
    trend_s  = f"{'▲' if trend >= 0 else '▼'} {abs(trend):.1%} vs avg"

    def card(title, value, cls="", sub=""):
        return html.Div([
            html.Div(str(value), className="summary-card-value",
                     style={"color": "#22c55e" if cls == "green" else
                            "#ef4444" if cls == "red" else
                            "#ff6b00" if cls == "amber" else
                            "#60a5fa"}),
            html.Div(title, className="summary-card-label"),
            html.Div(sub, style={"fontSize": "11px", "color": "#64748b",
                                 "marginTop": "3px"}) if sub else None,
        ], className=f"summary-card {cls}")

    roi_cls = "green" if cum_roi >= 0 else "red"
    return [
        card("Overall HR Hit Rate", f"{overall_rate:.1%}", "amber"),
        card("High-Conf Hit Rate",  f"{high_rate:.1%}",
             "green" if high_rate >= WARN_THRESHOLD else "red"),
        card("Last 7-Day Hit Rate", f"{last7:.1%}",
             "green" if trend >= 0 else "red", trend_s),
        card("Cumulative ROI",      f"{cum_roi:+.2f}u", roi_cls),
        card("Days Tracked",        days),
        card("Resolved Predictions",f"{total:,}"),
    ]


# ─── Rolling 7-day accuracy ───────────────────────────────────────────────────

def _rolling_accuracy_chart(df: pd.DataFrame):
    if df.empty:
        return _empty_fig("No resolved predictions yet")

    daily = (
        df.groupby(["Date", "Confidence"])
        .agg(hits=("Hit_num", "sum"), total=("Hit_num", "count"))
        .reset_index()
    )
    daily["hit_rate"] = daily["hits"] / daily["total"]

    fig = go.Figure()
    for conf in ["High", "Medium", "Low"]:
        sub = daily[daily["Confidence"] == conf].sort_values("Date")
        if sub.empty:
            continue

        # Rolling 7-day window (min 3 data points so early days aren't noisy)
        sub = sub.set_index("Date").asfreq("D").ffill()
        roll = sub["hit_rate"].rolling(7, min_periods=3).mean().reset_index()
        roll.columns = ["Date", "roll_rate"]

        fig.add_trace(go.Scatter(
            x=roll["Date"], y=roll["roll_rate"],
            mode="lines", name=f"{conf} (7-day)",
            line={"color": CONF_COLORS[conf], "width": 2.5},
        ))
        # Daily dots (faint)
        fig.add_trace(go.Scatter(
            x=sub.index, y=sub["hit_rate"],
            mode="markers", name=f"{conf} daily",
            marker={"color": CONF_COLORS[conf], "size": 4, "opacity": .4},
            showlegend=False,
        ))

    # 15 % warning line
    fig.add_hline(
        y=WARN_THRESHOLD, line_dash="dot",
        line_color="#ef4444", line_width=1.5,
        annotation_text="15% threshold",
        annotation_position="bottom right",
        annotation_font={"color": "#ef4444", "size": 11},
    )
    layout = _chart_layout(h=300)
    layout["yaxis"].update({"tickformat": ".0%", "title": "Hit Rate"})
    layout["xaxis"].update({"title": "Date"})
    fig.update_layout(**layout)
    return fig


# ─── Confidence calibration ───────────────────────────────────────────────────

def _calibration_chart(df: pd.DataFrame):
    if df.empty:
        return _empty_fig("No resolved predictions yet")

    # Exclude stub rows added for unmatched HR hitters (no model prediction)
    df = df[df["Confidence"].isin(CONF_COLORS)].copy()
    if df.empty:
        return _empty_fig("No resolved model predictions yet")

    summary = (
        df.groupby("Confidence")
        .agg(
            total=("Hit_num", "count"),
            hits=("Hit_num", "sum"),
            avg_prob=("HR_Probability", "mean"),
        )
        .reset_index()
    )
    summary["hit_rate"]  = summary["hits"] / summary["total"]
    summary["color"]     = summary["Confidence"].map(CONF_COLORS)

    # Sort High → Medium → Low
    order = {"High": 0, "Medium": 1, "Low": 2}
    summary["_ord"] = summary["Confidence"].map(order)
    summary = summary.sort_values("_ord")

    fig = go.Figure()

    # Actual hit rate bars
    fig.add_trace(go.Bar(
        x=summary["Confidence"], y=summary["hit_rate"],
        name="Actual Hit Rate",
        marker_color=summary["color"].tolist(),
        text=summary["hit_rate"].apply(lambda v: f"{v:.1%}"),
        textposition="outside",
        width=0.35,
        offset=-0.2,
    ))

    # Predicted avg probability bars (lighter shade)
    fig.add_trace(go.Bar(
        x=summary["Confidence"], y=summary["avg_prob"],
        name="Avg Predicted Prob",
        marker_color=[
            "rgba({},{},{},0.38)".format(int(c[1:3], 16), int(c[3:5], 16), int(c[5:7], 16))
            if isinstance(c, str) and c.startswith("#") and len(c) >= 7
            else "rgba(148,163,184,0.38)"
            for c in summary["color"].tolist()
        ],
        text=summary["avg_prob"].apply(lambda v: f"{v:.1%}"),
        textposition="outside",
        width=0.35,
        offset=0.2,
    ))

    # Sample-size annotation
    for _, row in summary.iterrows():
        fig.add_annotation(
            x=row["Confidence"], y=-0.04,
            text=f"n={int(row['total'])}",
            showarrow=False, xref="x", yref="paper",
            font={"size": 11, "color": "#8e909c"},
        )

    fig.add_hline(
        y=WARN_THRESHOLD, line_dash="dot",
        line_color="#ef4444", line_width=1.5,
    )
    layout = _chart_layout(h=300)
    layout["yaxis"].update({"tickformat": ".0%", "title": "Rate", "range": [0, None]})
    layout["xaxis"].update({"title": "Confidence Tier"})
    layout["legend"].update({"orientation": "h", "y": -0.2})
    fig.update_layout(**layout, barmode="overlay")
    return fig


# ─── Cumulative accuracy chart ────────────────────────────────────────────────

def _accuracy_chart(df: pd.DataFrame):
    if df.empty or "Date" not in df.columns:
        return _empty_fig("No accuracy data yet")

    df = df.sort_values("Date")
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=df["Date"], y=df["Cumulative_Accuracy"],
        mode="lines", name="Cumulative",
        line={"color": "#3b82f6", "width": 2.5},
        fill="tozeroy", fillcolor="rgba(59,130,246,.12)",
    ))
    fig.add_trace(go.Scatter(
        x=df["Date"], y=df["Daily_Accuracy"],
        mode="lines", name="Daily",
        line={"color": "#93c5fd", "width": 1, "dash": "dot"},
    ))
    fig.add_hline(
        y=WARN_THRESHOLD, line_dash="dot",
        line_color="#ef4444", line_width=1.5,
    )
    layout = _chart_layout(h=280)
    layout["yaxis"].update({"tickformat": ".1%"})
    fig.update_layout(**layout)
    return fig


# ─── ROI chart ────────────────────────────────────────────────────────────────

def _roi_chart(df: pd.DataFrame):
    if df.empty or "ROI_Flat_Bet" not in df.columns:
        return _empty_fig("No ROI data yet")

    df = df.sort_values("Date")
    cum_roi = df["ROI_Flat_Bet"].cumsum()
    final   = float(cum_roi.iloc[-1]) if len(cum_roi) else 0
    color   = "#22c55e" if final >= 0 else "#ef4444"

    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=df["Date"], y=cum_roi,
        mode="lines+markers", fill="tozeroy",
        line={"color": color, "width": 2.5},
        fillcolor=f"{'rgba(34,197,94,.1)' if final >= 0 else 'rgba(239,68,68,.1)'}",
        name="Cumulative ROI",
        marker={"size": 5},
    ))
    fig.add_hline(y=0, line_dash="dash", line_color="#8e909c", line_width=1)
    layout = _chart_layout(h=280)
    layout["yaxis"].update({"title": "Units (u)"})
    fig.update_layout(**layout)
    return fig


# ─── Feature importance ───────────────────────────────────────────────────────

def _feature_importance_chart(df: pd.DataFrame):
    if df.empty:
        return _empty_fig("Train the model to see feature importances")

    latest = (
        df.sort_values("Date")
        .groupby("Feature")["Importance_Score"]
        .last()
        .reset_index()
        .sort_values("Importance_Score", ascending=True)
        .tail(15)
    )

    colors = [
        "#ff6b00" if s == latest["Importance_Score"].max()
        else "#3b82f6" if s >= latest["Importance_Score"].quantile(.75)
        else "#1e3566"
        for s in latest["Importance_Score"]
    ]

    fig = go.Figure(go.Bar(
        x=latest["Importance_Score"],
        y=latest["Feature"],
        orientation="h",
        marker_color=colors,
        text=latest["Importance_Score"].apply(lambda v: f"{v:.3f}"),
        textposition="outside",
    ))
    layout = _chart_layout(h=420)
    layout["yaxis"].update({"tickfont": {"size": 12}})
    fig.update_layout(**layout, xaxis_title="SHAP Importance Score")
    return fig


# ─── Shared helpers ───────────────────────────────────────────────────────────

def _chart_layout(h: int = 280) -> dict:
    return {
        "paper_bgcolor": "#151515",
        "plot_bgcolor":  "#101416",
        "height":        h,
        "margin":        dict(l=60, r=30, t=10, b=50),
        "legend":        {"orientation": "h", "y": -0.2,
                          "font": {"color": "#8e909c"}},
        "font":          {"family": "Manrope, Inter, sans-serif", "size": 12, "color": "#f2f2f2"},
        "xaxis":         {"gridcolor": "rgba(255,255,255,.06)", "color": "#8e909c",
                          "zerolinecolor": "rgba(255,255,255,.06)"},
        "yaxis":         {"gridcolor": "rgba(255,255,255,.06)", "color": "#8e909c",
                          "zerolinecolor": "rgba(255,255,255,.06)"},
    }


def _empty_fig(msg: str):
    fig = go.Figure()
    fig.add_annotation(
        text=msg, xref="paper", yref="paper",
        x=0.5, y=0.5, showarrow=False,
        font={"size": 14, "color": "#8e909c"},
    )
    fig.update_layout(
        height=280, paper_bgcolor="#151515", plot_bgcolor="#101416",
        margin=dict(l=40, r=20, t=20, b=40),
    )
    return fig

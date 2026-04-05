"""
Player Stats page.
Per-player season performance: hit rate, calibration bias, last-30-day form.
Data comes from the Player_Stats sheet rebuilt after each results update.
"""

from pathlib import Path

import pandas as pd
import plotly.graph_objects as go
from dash import Input, Output, callback, dash_table, dcc, html
from openpyxl import load_workbook

EXCEL_PATH = Path("MLB_HR_Predictions.xlsx")

PLAYER_COLUMNS = [
    "Player", "Team",
    "Season_Predictions", "Season_HRs_Actual", "Season_Hit_Rate",
    "Avg_HR_Prob", "Calibration_Bias",
    "Last_30_Predictions", "Last_30_HRs", "Last_30_Hit_Rate",
    "Last_Updated",
]

CONF_COLORS = {"High": "#22c55e", "Medium": "#ff6b00", "Low": "#8e909c"}
_GREEN = "#22c55e"
_RED   = "#ef4444"
_ORANGE = "#ff6b00"
_GRAY   = "#8e909c"


# ─── Layout ───────────────────────────────────────────────────────────────────

def layout():
    return html.Div([
        html.Div([
            html.Div([
                html.H2("Player Performance Tracker", className="page-title"),
                html.P("Per-player hit rates, model calibration, and recent form",
                       className="page-subtitle"),
            ]),
        ], className="page-header"),

        # KPI strip
        html.Div(id="ps-kpi-strip", className="summary-cards"),

        # ── Row 1: Top/Bottom performers ──────────────────────────────────────
        html.Div([
            html.Div([
                html.H3("Top 10 Hit-Rate Players (season)", className="section-title"),
                html.Span("Players the model has been most accurate on",
                          style={"fontSize": "12px", "color": "#8e909c",
                                 "marginBottom": "10px", "display": "block"}),
                dcc.Graph(id="ps-top-players-chart", config={"displayModeBar": False}),
            ], className="chart-section", style={"flex": "1"}),

            html.Div([
                html.H3("Calibration: Predicted vs Actual Hit Rate",
                        className="section-title"),
                html.Span("Dots above the diagonal → model underestimates that player",
                          style={"fontSize": "12px", "color": "#8e909c",
                                 "marginBottom": "10px", "display": "block"}),
                dcc.Graph(id="ps-calibration-scatter", config={"displayModeBar": False}),
            ], className="chart-section", style={"flex": "1"}),
        ], style={"display": "flex", "gap": "16px", "flexWrap": "wrap"}),

        # ── Row 2: Last-30-day form ────────────────────────────────────────────
        html.Div([
            html.H3("Last-30-Day Form (min 3 predictions)", className="section-title"),
            html.Span("Recent performance — recency bias correction is applied here",
                      style={"fontSize": "12px", "color": "#8e909c",
                             "marginBottom": "10px", "display": "block"}),
            dcc.Graph(id="ps-recent-form-chart", config={"displayModeBar": False}),
        ], className="chart-section"),

        # ── Row 3: Full player table ───────────────────────────────────────────
        html.Div([
            html.H3("Full Season Player Table", className="section-title"),
            html.Div(id="ps-table-container"),
        ], className="chart-section"),

        dcc.Interval(id="ps-refresh", interval=10 * 60 * 1000, n_intervals=0),
    ], className="page-container")


# ─── Data loader ──────────────────────────────────────────────────────────────

def _load_player_stats() -> pd.DataFrame:
    if not EXCEL_PATH.exists():
        return pd.DataFrame(columns=PLAYER_COLUMNS)
    try:
        wb  = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        if "Player_Stats" not in wb.sheetnames:
            wb.close()
            return pd.DataFrame(columns=PLAYER_COLUMNS)
        ws   = wb["Player_Stats"]
        data = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
    except Exception:
        return pd.DataFrame(columns=PLAYER_COLUMNS)

    if not data:
        return pd.DataFrame(columns=PLAYER_COLUMNS)

    row_len = len(data[0])
    cols    = PLAYER_COLUMNS[:row_len] if row_len <= len(PLAYER_COLUMNS) else PLAYER_COLUMNS
    df      = pd.DataFrame(data, columns=cols)
    for col in PLAYER_COLUMNS:
        if col not in df.columns:
            df[col] = None

    df = df[df["Player"].notna()].copy()
    for num_col in ["Season_Predictions", "Season_HRs_Actual", "Season_Hit_Rate",
                    "Avg_HR_Prob", "Calibration_Bias",
                    "Last_30_Predictions", "Last_30_HRs", "Last_30_Hit_Rate"]:
        df[num_col] = pd.to_numeric(df[num_col], errors="coerce").fillna(0)
    return df


# ─── Callbacks ────────────────────────────────────────────────────────────────

@callback(
    Output("ps-kpi-strip",          "children"),
    Output("ps-top-players-chart",  "figure"),
    Output("ps-calibration-scatter","figure"),
    Output("ps-recent-form-chart",  "figure"),
    Output("ps-table-container",    "children"),
    Input("ps-refresh", "n_intervals"),
)
def update_player_page(n_intervals):
    df = _load_player_stats()

    if df.empty:
        empty_fig = go.Figure().update_layout(
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            xaxis={"visible": False}, yaxis={"visible": False},
            annotations=[{"text": "No resolved predictions yet",
                          "showarrow": False, "font": {"color": "#8e909c", "size": 14}}],
        )
        kpis = html.P("No player data yet.", style={"color": "#8e909c"})
        return kpis, empty_fig, empty_fig, empty_fig, html.P("No data.")

    # ── KPI strip ─────────────────────────────────────────────────────────────
    total_players = len(df)
    tracked       = int((df["Season_Predictions"] >= 5).sum())
    avg_hit_rate  = df.loc[df["Season_Predictions"] >= 5, "Season_Hit_Rate"].mean()
    well_calibrated = int((df["Calibration_Bias"].abs() < 0.03).sum())

    kpis = html.Div([
        _kpi("Players Tracked",      str(total_players),               "person"),
        _kpi("With ≥5 Predictions",  str(tracked),                     "bar_chart"),
        _kpi("Avg Hit Rate",         f"{avg_hit_rate:.1%}" if tracked else "—", "sports_baseball"),
        _kpi("Well-Calibrated",      str(well_calibrated),             "verified"),
    ], className="summary-cards")

    # ── Top 10 chart ──────────────────────────────────────────────────────────
    top10 = (
        df[df["Season_Predictions"] >= 5]
        .nlargest(10, "Season_Hit_Rate")
        .sort_values("Season_Hit_Rate")
    )
    bar_colors = [
        _GREEN if r >= 0.30 else _ORANGE if r >= 0.15 else _GRAY
        for r in top10["Season_Hit_Rate"]
    ]
    top_fig = go.Figure(go.Bar(
        x=top10["Season_Hit_Rate"],
        y=top10["Player"],
        orientation="h",
        marker_color=bar_colors,
        text=[f"{r:.0%}" for r in top10["Season_Hit_Rate"]],
        textposition="outside",
        customdata=top10[["Team", "Season_Predictions", "Season_HRs_Actual"]].values,
        hovertemplate=(
            "<b>%{y}</b> (%{customdata[0]})<br>"
            "Hit Rate: %{x:.1%}<br>"
            "Predictions: %{customdata[1]}<br>"
            "Season HRs: %{customdata[2]}<extra></extra>"
        ),
    ))
    top_fig.update_layout(**_chart_layout(), margin={"l": 140, "r": 60, "t": 10, "b": 40})
    top_fig.update_xaxes(tickformat=".0%", range=[0, min(top10["Season_Hit_Rate"].max() * 1.3, 1.0)])

    # ── Calibration scatter ───────────────────────────────────────────────────
    cal_df = df[df["Season_Predictions"] >= 5].copy()
    dot_colors = [
        _GREEN if abs(b) < 0.03 else _RED if b > 0.05 else _ORANGE
        for b in cal_df["Calibration_Bias"]
    ]
    cal_fig = go.Figure()
    # Diagonal reference line (perfect calibration)
    max_prob = cal_df["Avg_HR_Prob"].max() if not cal_df.empty else 0.3
    cal_fig.add_trace(go.Scatter(
        x=[0, max_prob * 1.1], y=[0, max_prob * 1.1],
        mode="lines",
        line={"color": "rgba(255,255,255,.15)", "dash": "dot"},
        showlegend=False,
        hoverinfo="skip",
    ))
    cal_fig.add_trace(go.Scatter(
        x=cal_df["Avg_HR_Prob"],
        y=cal_df["Season_Hit_Rate"],
        mode="markers",
        marker={
            "color": dot_colors,
            "size": [max(6, min(14, p / 5)) for p in cal_df["Season_Predictions"]],
            "opacity": 0.85,
            "line": {"width": 1, "color": "rgba(255,255,255,.2)"},
        },
        text=cal_df["Player"],
        customdata=cal_df[["Team", "Season_Predictions", "Calibration_Bias"]].values,
        hovertemplate=(
            "<b>%{text}</b> (%{customdata[0]})<br>"
            "Avg Predicted Prob: %{x:.1%}<br>"
            "Actual Hit Rate:    %{y:.1%}<br>"
            "Bias: %{customdata[2]:+.3f}<br>"
            "n=%{customdata[1]}<extra></extra>"
        ),
        showlegend=False,
    ))
    cal_fig.update_layout(**_chart_layout(), margin={"l": 50, "r": 20, "t": 10, "b": 50})
    cal_fig.update_xaxes(tickformat=".0%", title_text="Avg Predicted Probability")
    cal_fig.update_yaxes(tickformat=".0%", title_text="Actual Hit Rate")

    # ── Last-30-day form chart ────────────────────────────────────────────────
    recent = (
        df[df["Last_30_Predictions"] >= 3]
        .nlargest(20, "Last_30_Hit_Rate")
        .sort_values("Last_30_Hit_Rate")
    )
    r_colors = [
        _GREEN if r >= 0.30 else _ORANGE if r >= 0.15 else _GRAY
        for r in recent["Last_30_Hit_Rate"]
    ]
    recent_fig = go.Figure(go.Bar(
        x=recent["Last_30_Hit_Rate"],
        y=recent["Player"],
        orientation="h",
        marker_color=r_colors,
        text=[f"{r:.0%}" for r in recent["Last_30_Hit_Rate"]],
        textposition="outside",
        customdata=recent[["Last_30_Predictions", "Last_30_HRs"]].values,
        hovertemplate=(
            "<b>%{y}</b><br>"
            "Last-30 Hit Rate: %{x:.1%}<br>"
            "Predictions: %{customdata[0]}<br>"
            "HRs Hit: %{customdata[1]}<extra></extra>"
        ),
    ))
    recent_fig.update_layout(
        **_chart_layout(),
        height=max(300, len(recent) * 28),
        margin={"l": 140, "r": 60, "t": 10, "b": 40},
    )
    recent_fig.update_xaxes(tickformat=".0%")

    # ── Full table ────────────────────────────────────────────────────────────
    table_df = (
        df[df["Season_Predictions"] >= 1]
        .sort_values("Season_Hit_Rate", ascending=False)
        [[
            "Player", "Team",
            "Season_Predictions", "Season_HRs_Actual", "Season_Hit_Rate",
            "Avg_HR_Prob", "Calibration_Bias",
            "Last_30_Predictions", "Last_30_Hit_Rate",
        ]]
        .copy()
    )
    for pct_col in ["Season_Hit_Rate", "Avg_HR_Prob", "Calibration_Bias", "Last_30_Hit_Rate"]:
        table_df[pct_col] = table_df[pct_col].apply(lambda v: f"{v:.1%}" if pd.notna(v) else "—")

    table = dash_table.DataTable(
        data=table_df.to_dict("records"),
        columns=[{"name": c.replace("_", " "), "id": c} for c in table_df.columns],
        style_table={"overflowX": "auto"},
        style_cell={
            "backgroundColor": "var(--surface)",
            "color":           "var(--on-surface)",
            "border":          "1px solid var(--border)",
            "fontFamily":      "'Inter', sans-serif",
            "fontSize":        "13px",
            "padding":         "8px 12px",
            "textAlign":       "left",
        },
        style_header={
            "backgroundColor": "var(--surface-variant)",
            "color":           "var(--on-surface)",
            "fontWeight":      "700",
            "fontSize":        "11px",
            "textTransform":   "uppercase",
            "letterSpacing":   "0.8px",
        },
        style_data_conditional=[
            {"if": {"row_index": "odd"}, "backgroundColor": "rgba(255,255,255,.02)"},
        ],
        page_size=25,
        sort_action="native",
        filter_action="native",
    )

    return kpis, top_fig, cal_fig, recent_fig, table


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _kpi(label: str, value: str, icon: str):
    return html.Div([
        html.Div([
            html.Span(icon, className="material-symbols-outlined",
                      style={"fontSize": "20px", "color": "var(--primary)"}),
            html.Span(label, className="metric-label"),
        ], style={"display": "flex", "alignItems": "center", "gap": "8px"}),
        html.Div(value, className="metric-value"),
    ], className="metric-card")


def _chart_layout() -> dict:
    return dict(
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font={"family": "'Inter', sans-serif", "color": "#e2e8f0", "size": 12},
        xaxis={"gridcolor": "rgba(255,255,255,.06)", "showgrid": True},
        yaxis={"gridcolor": "rgba(255,255,255,.06)", "showgrid": False},
        hoverlabel={"bgcolor": "#1e1e2e", "font_size": 12},
    )

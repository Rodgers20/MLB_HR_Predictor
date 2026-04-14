"""
Today's Predictions page.
Reads predictions from MLB_HR_Predictions.xlsx (written by the morning scheduler).
Kinetic Analytics design: 3-column top picks, hero spotlight, venue card, full table.
"""

from datetime import date
from pathlib import Path
from urllib.parse import quote_plus

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from dash import Input, Output, State, callback, ctx, dash_table, dcc, html
from openpyxl import load_workbook

EXCEL_PATH = Path("MLB_HR_Predictions.xlsx")
PRED_COLUMNS = [
    "Date", "Player", "Team", "Opponent", "Pitcher",
    "HR_Probability", "Confidence", "Park_Factor",
    "Temp_F", "Wind_Speed_MPH", "Wind_Direction", "Is_Indoor",
    "Home_Game", "Predicted_HRs", "Actual_HRs", "Hit",
    "Insight_Text",   # SHAP explanation (v2+; absent in old Excel files)
]
CONF_COLORS = {"High": "#22c55e", "Medium": "#ff6b00", "Low": "#8e909c"}


def layout():
    today_str = date.today().strftime("%B %d, %Y")
    return html.Div([
        # ── Page header ──────────────────────────────────────────────────────
        html.Div([
            html.Div([
                html.H2("Today's Top HR Predictions", className="page-title"),
                html.P("Algorithmic probability spikes for today's matchups",
                       className="page-subtitle"),
            ]),
            html.Div([
                html.Span("Live Processing", style={
                    "fontSize": "9px", "fontWeight": "800",
                    "textTransform": "uppercase", "letterSpacing": "2px",
                    "color": "var(--primary)", "display": "block",
                    "textAlign": "right",
                }),
                html.Div([
                    html.Span(style={
                        "width": "8px", "height": "8px", "borderRadius": "50%",
                        "background": "var(--primary)", "flexShrink": "0",
                    }),
                    html.Span(today_str, style={
                        "fontSize": "12px", "fontWeight": "600",
                        "color": "var(--on-surface-variant)",
                    }),
                ], style={"display": "flex", "alignItems": "center", "gap": "8px", "marginTop": "4px"}),
            ], style={"textAlign": "right"}),
        ], className="page-header"),

        # ── Filters ──────────────────────────────────────────────────────────
        html.Div([
            html.Div([
                html.Label("Confidence"),
                dcc.Dropdown(
                    id="conf-filter",
                    options=[
                        {"label": "All", "value": "All"},
                        {"label": "High Only", "value": "High"},
                        {"label": "High + Medium", "value": "High+Medium"},
                    ],
                    value="All", clearable=False,
                    style={"width": "190px", "fontSize": "13px"},
                ),
            ]),
            html.Div([
                html.Label("Team"),
                dcc.Dropdown(
                    id="team-filter",
                    options=[{"label": "All Teams", "value": "All"}],
                    value="All", clearable=False,
                    style={"width": "175px", "fontSize": "13px"},
                ),
            ]),
            html.Div([
                html.Label("Wind"),
                dcc.Dropdown(
                    id="wind-filter",
                    options=[
                        {"label": "All", "value": "All"},
                        {"label": "Out (boost)",     "value": "out"},
                        {"label": "Calm",             "value": "calm"},
                        {"label": "Cross",            "value": "cross"},
                        {"label": "In (suppressor)", "value": "in"},
                    ],
                    value="All", clearable=False,
                    style={"width": "175px", "fontSize": "13px"},
                ),
            ]),
            html.Div([
                html.Label("Sort By"),
                dcc.Dropdown(
                    id="sort-by",
                    options=[
                        {"label": "HR Probability", "value": "probability"},
                        {"label": "Team",           "value": "team"},
                    ],
                    value="probability", clearable=False,
                    style={"width": "175px", "fontSize": "13px"},
                ),
            ]),
        ], className="filters-row"),

        # ── Weather banner ────────────────────────────────────────────────────
        html.Div(id="weather-banner"),

        # ── Summary cards ─────────────────────────────────────────────────────
        html.Div(id="summary-cards", className="summary-cards"),

        # ── Top 3 picks grid (3-column vertical cards) ────────────────────────
        html.Div(id="picks-grid"),

        # ── Hero bento: spotlight (2/3) + venue card (1/3) ───────────────────
        html.Div([
            html.Div(id="hero-spotlight-container"),
            html.Div(id="venue-card-container"),
        ], className="hero-bento"),

        # ── Daily 3-leg parlay ────────────────────────────────────────────────
        html.Div(id="daily-parlay"),

        # ── Featured 10 players (ranks 4–13) ─────────────────────────────────
        html.Div(id="featured-players-section"),

        # ── Full ranked table ─────────────────────────────────────────────────
        html.Div([
            html.Div([
                html.Div([
                    html.H4("Full Prediction List", style={
                        "margin": 0, "fontSize": "15px",
                        "fontWeight": "800", "color": "var(--primary)",
                    }),
                    html.P("Sortable · ranked by HR probability",
                           style={"margin": "2px 0 0", "fontSize": "12px",
                                  "color": "var(--on-surface-variant)"}),
                ]),
            ], className="table-title-bar"),
            dash_table.DataTable(
                id="predictions-table",
                columns=[
                    {"name": "#",           "id": "rank",          "type": "numeric"},
                    {"name": "Player",      "id": "Player"},
                    {"name": "Team",        "id": "Team"},
                    {"name": "Opponent",    "id": "Opponent"},
                    {"name": "vs Pitcher",  "id": "Pitcher"},
                    {"name": "HR Prob",     "id": "HR_Probability", "type": "numeric",
                     "format": {"specifier": ".1%"}},
                    {"name": "Confidence",  "id": "Confidence"},
                    {"name": "Park",        "id": "Park_Factor",   "type": "numeric"},
                    {"name": "Temp °F",     "id": "Temp_F",        "type": "numeric"},
                    {"name": "Wind",        "id": "wind_display"},
                    {"name": "Home?",       "id": "home_display"},
                    {"name": "Deep Dive",   "id": "deep_dive_link",
                     "presentation": "markdown"},
                ],
                data=[],
                sort_action="native",
                filter_action="native",
                page_size=30,
                style_table={"overflowX": "auto"},
                style_header={
                    "backgroundColor": "#1c1c1c",
                    "color": "#f2f2f2",
                    "fontWeight": "700",
                    "textAlign": "center",
                    "padding": "11px 10px",
                    "fontSize": "10px",
                    "letterSpacing": "1.5px",
                    "textTransform": "uppercase",
                    "borderBottom": "2px solid #ff6b00",
                    "border": "none",
                },
                style_cell={
                    "textAlign": "center",
                    "padding": "14px 16px",
                    "fontFamily": "Manrope, sans-serif",
                    "fontSize": "13px",
                    "backgroundColor": "#131313",
                    "color": "#f0dfd8",
                    "border": "none",
                },
                style_cell_conditional=[
                    {"if": {"column_id": "Player"},         "textAlign": "left", "fontWeight": "600"},
                    {"if": {"column_id": "rank"},           "width": "48px"},
                    {"if": {"column_id": "deep_dive_link"}, "width": "80px", "textAlign": "center",
                     "color": "#ff6b00", "fontWeight": "600", "fontSize": "12px"},
                ],
                style_data_conditional=[
                    {"if": {"filter_query": '{Confidence} = "High"'},
                     "backgroundColor": "rgba(34,197,94,.07)", "color": "#86efac"},
                    {"if": {"filter_query": '{Confidence} = "Medium"'},
                     "backgroundColor": "rgba(255,107,0,.07)", "color": "#fdba74"},
                    {"if": {"row_index": "odd"},
                     "backgroundColor": "#1a1a1a"},
                ],
            ),
        ], className="table-container"),

        # ── Probability distribution ──────────────────────────────────────────
        html.Div([
            html.H3("HR Probability Distribution", className="section-title"),
            dcc.Graph(id="prob-distribution-chart", config={"displayModeBar": False}),
        ], className="chart-section"),

        dcc.Store(id="predictions-store"),
        dcc.Interval(id="refresh-interval", interval=5 * 60 * 1000, n_intervals=0),

        # ── Player stat modal ─────────────────────────────────────────────────
        html.Div([
            html.Div([
                html.Button("✕", id="modal-close-btn", className="modal-close-btn"),
                html.Div(id="player-modal-content"),
            ], className="player-modal"),
        ], id="player-modal-overlay", className="player-modal-overlay",
           style={"display": "none"}),
    ], className="page-container")


# ─── Data loader ──────────────────────────────────────────────────────────────

def _load_todays_predictions() -> pd.DataFrame:
    if not EXCEL_PATH.exists():
        return pd.DataFrame()
    try:
        wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        ws = wb["Predictions"]
        data = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
    except Exception:
        return pd.DataFrame()

    if not data:
        return pd.DataFrame()

    # Gracefully handle old Excel files that don't have Insight_Text column
    row_len = len(data[0]) if data else 0
    cols = PRED_COLUMNS[:row_len] if row_len <= len(PRED_COLUMNS) else PRED_COLUMNS
    df = pd.DataFrame(data, columns=cols)
    if "Insight_Text" not in df.columns:
        df["Insight_Text"] = ""

    today_str = date.today().isoformat()
    df = df[df["Date"] == today_str].copy()
    if df.empty:
        return df

    df["HR_Probability"] = pd.to_numeric(df["HR_Probability"], errors="coerce")
    df = (
        df.sort_values("HR_Probability", ascending=False)
          .drop_duplicates(subset=["Player", "Team", "Pitcher"], keep="first")
          .reset_index(drop=True)
    )
    return df


# ─── Callbacks ────────────────────────────────────────────────────────────────

@callback(
    Output("predictions-store", "data"),
    Output("team-filter", "options"),
    Input("refresh-interval", "n_intervals"),
)
def load_predictions(_n):
    df = _load_todays_predictions()
    if df.empty:
        return [], [{"label": "All Teams", "value": "All"}]
    teams = [{"label": "All Teams", "value": "All"}] + [
        {"label": t, "value": t} for t in sorted(df["Team"].dropna().unique())
    ]
    return df.to_dict("records"), teams


@callback(
    Output("predictions-table", "data"),
    Output("summary-cards", "children"),
    Output("prob-distribution-chart", "figure"),
    Output("picks-grid", "children"),
    Output("daily-parlay", "children"),
    Output("weather-banner", "children"),
    Output("hero-spotlight-container", "children"),
    Output("venue-card-container", "children"),
    Output("featured-players-section", "children"),
    Input("predictions-store", "data"),
    Input("conf-filter", "value"),
    Input("team-filter", "value"),
    Input("wind-filter", "value"),
    Input("sort-by", "value"),
)
def update_table(data, conf_filter, team_filter, wind_filter, sort_by):
    empty = (
        [], _empty_cards(), _empty_chart(), _empty_picks(),
        [], [], _empty_hero(), _empty_venue(), [],
    )
    if not data:
        return empty

    df = pd.DataFrame(data)
    df["HR_Probability"] = pd.to_numeric(df["HR_Probability"], errors="coerce")

    if conf_filter == "High":
        df = df[df["Confidence"] == "High"]
    elif conf_filter == "High+Medium":
        df = df[df["Confidence"].isin(["High", "Medium"])]
    if team_filter != "All":
        df = df[df["Team"] == team_filter]
    if wind_filter != "All":
        df = df[df["Wind_Direction"] == wind_filter]

    if sort_by == "team":
        df = df.sort_values(["Team", "HR_Probability"], ascending=[True, False]).reset_index(drop=True)
    else:
        df = df.sort_values("HR_Probability", ascending=False).reset_index(drop=True)

    df["rank"] = df.index + 1
    df["wind_display"] = (
        df["Wind_Speed_MPH"].fillna(0).astype(int).astype(str) + " mph " +
        df["Wind_Direction"].fillna("calm")
    )
    df["home_display"] = df["Home_Game"].map(
        {True: "Home", False: "Away", 1: "Home", 0: "Away"}
    )

    df["deep_dive_link"] = df.apply(
        lambda r: f"[↗ Stats](/player?name={quote_plus(str(r['Player']))}&pitcher={quote_plus(str(r.get('Pitcher', '')))})",
        axis=1,
    )
    table_data = df[[
        "rank", "Player", "Team", "Opponent", "Pitcher", "HR_Probability",
        "Confidence", "Park_Factor", "Temp_F", "wind_display", "home_display",
        "deep_dive_link",
    ]].to_dict("records")

    deduped   = df.drop_duplicates(subset=["Player"], keep="first")
    picks     = _build_top_picks_grid(deduped.head(3))
    featured  = _build_featured_players(deduped.iloc[3:13])
    parlay    = _build_parlay_card(deduped)
    weather   = _build_weather_banner(df)
    hero      = _build_hero_spotlight(deduped)
    venue     = _build_venue_card(df)

    return (
        table_data, _build_cards(df), _build_dist_chart(df),
        picks, parlay, weather, hero, venue, featured,
    )


# ─── Top 3 picks vertical cards ──────────────────────────────────────────────

def _build_top_picks_grid(df: pd.DataFrame):
    if df.empty:
        return _empty_picks()

    try:
        from utils.player_photos import batch_headshot_urls
        photo_map = batch_headshot_urls(df["Player"].tolist(), width=80)
    except Exception:
        photo_map = {}

    cards = []
    for rank_offset, (_, row) in enumerate(df.iterrows()):
        prob    = float(row.get("HR_Probability", 0) or 0)
        name    = str(row.get("Player", ""))
        team    = str(row.get("Team", ""))
        opp     = str(row.get("Opponent", ""))
        pitcher = str(row.get("Pitcher", "") or "")
        conf    = str(row.get("Confidence", "") or "")
        insight = str(row.get("Insight_Text", "") or "")
        photo   = photo_map.get(name, "")
        matchup = f"{team} vs {opp}" if opp else team
        bar_w   = f"{min(prob * 250, 100):.1f}%"  # scale so 40% fills 100%
        rank_num = rank_offset + 1

        # Parse insight into bullet reasons
        reason_bullets = _parse_insight_bullets(insight)

        conf_color = {"High": "#22c55e", "Medium": "#ff6b00", "Low": "#8e909c"}.get(conf, "#8e909c")

        card = html.Div([
            html.Div([
                html.Div([
                    html.Span(f"#{rank_num}", style={
                        "position": "absolute", "top": "10px", "left": "10px",
                        "fontSize": "11px", "fontWeight": "900",
                        "color": "#ff6b00", "letterSpacing": "1px",
                    }),
                    html.Img(
                        src=photo, alt=name, className="top-pick-photo",
                        style={"background": "var(--surface-highest)"},
                    ) if photo else html.Div(
                        name[:2].upper(),
                        className="top-pick-photo",
                        style={
                            "display": "flex", "alignItems": "center",
                            "justifyContent": "center", "fontSize": "18px",
                            "fontWeight": "900", "color": "var(--primary)",
                            "background": "var(--surface-highest)",
                        },
                    ),
                    html.Div([
                        html.Span("HR Prob", className="top-pick-prob-label"),
                        html.Div(f"{prob:.1%}", className="top-pick-prob-value"),
                    ], className="top-pick-prob-block"),
                ], className="top-pick-card-header", style={"position": "relative"}),

                html.Div([
                    dcc.Link(
                        name,
                        href=f"/player?name={quote_plus(name)}&pitcher={quote_plus(pitcher)}",
                        className="top-pick-name",
                        style={"textDecoration": "none", "cursor": "pointer",
                               "color": "inherit", "display": "block"},
                    ),
                    html.P(matchup, className="top-pick-matchup"),
                    html.Span(conf, style={
                        "fontSize": "10px", "fontWeight": "700",
                        "color": conf_color, "textTransform": "uppercase",
                        "letterSpacing": "1px",
                    }),
                ]),

                # ── Why the model picked this player ──────────────────────
                html.Div([
                    html.Div("Why this pick", style={
                        "fontSize": "10px", "fontWeight": "700",
                        "textTransform": "uppercase", "letterSpacing": "1.5px",
                        "color": "#64748b", "marginBottom": "6px",
                    }),
                    html.Div([
                        html.Div([
                            html.Span("▸ ", style={"color": "#ff6b00"}),
                            html.Span(r, style={"fontSize": "12px", "color": "#cbd5e1"}),
                        ], style={"marginBottom": "3px"})
                        for r in reason_bullets
                    ]) if reason_bullets else html.P(
                        "No detailed reasoning available.",
                        style={"fontSize": "12px", "color": "#64748b", "fontStyle": "italic"},
                    ),
                    html.Div([
                        html.Span("vs ", style={"fontSize": "11px", "color": "#64748b"}),
                        html.Span(pitcher, style={"fontSize": "11px", "color": "#94a3b8", "fontWeight": "600"}),
                    ], style={"marginTop": "6px"}) if pitcher and pitcher != "TBD" else None,
                ], style={
                    "marginTop": "10px", "padding": "10px 12px",
                    "background": "rgba(255,255,255,.04)",
                    "borderRadius": "8px", "borderLeft": "2px solid #ff6b00",
                }),

                html.Div([
                    html.Div(style={"width": bar_w}, className="top-pick-bar-fill"),
                ], className="top-pick-bar-track"),
            ], className="top-pick-card-inner"),
        ], className="top-pick-card")
        cards.append(card)

    return html.Div(cards, className="top-picks-grid")


def _parse_insight_bullets(insight: str) -> list:
    """Split 'HR/FB rate 33% · ISO 0.333 · fly-ball hitter' into a list of phrases."""
    if not insight or insight == "Model-selected pick":
        return []
    parts = [p.strip() for p in insight.split("·") if p.strip()]
    return parts[:3]  # cap at 3 reasons


# ─── Hero spotlight (player to watch) ────────────────────────────────────────

def _lookup_batting_stats(player_name: str) -> dict:
    """Return AVG, HR, SLG, EV, Barrel% — current season preferred, prior season fallback."""
    try:
        from utils.data_fetcher import fetch_fangraphs_batting

        def _find_player(fg) -> object:
            parts = player_name.split()
            if len(parts) >= 2:
                return fg[
                    fg["Name"].str.contains(parts[0], case=False, na=False) &
                    fg["Name"].str.contains(parts[-1], case=False, na=False)
                ]
            return fg[fg["Name"].str.contains(player_name, case=False, na=False)]

        current_year = date.today().year
        row = _find_player(fetch_fangraphs_batting(current_year))
        if row.empty:
            row = _find_player(fetch_fangraphs_batting(current_year - 1))
        if row.empty:
            return {}
        r = row.iloc[0]
        return {
            "avg": r.get("AVG"),
            "hr":  int(r.get("HR", 0) or 0),
            "slg": r.get("SLG"),
            "ev":  r.get("EV"),
            "barrel_pct": r.get("Barrel%"),
            "hr_fb": r.get("HR/FB"),
            "pa":  int(r.get("PA", 0) or 0),
        }
    except Exception:
        return {}


def _build_featured_players(df: pd.DataFrame):
    """Build a grid of 10 featured player cards (ranks 4–13)."""
    if df.empty:
        return []

    try:
        from utils.player_photos import batch_headshot_urls
        photo_map = batch_headshot_urls(df["Player"].tolist(), width=60)
    except Exception:
        photo_map = {}

    cards = []
    for rank_offset, (_, row) in enumerate(df.iterrows()):
        prob    = float(row.get("HR_Probability", 0) or 0)
        name    = str(row.get("Player", ""))
        team    = str(row.get("Team", ""))
        opp     = str(row.get("Opponent", "") or "")
        pitcher = str(row.get("Pitcher", "") or "")
        conf    = str(row.get("Confidence", "") or "")
        insight = str(row.get("Insight_Text", "") or "")
        photo   = photo_map.get(name, "")
        rank_num = rank_offset + 4  # starts at 4

        reasons = _parse_insight_bullets(insight)
        conf_color = {"High": "#22c55e", "Medium": "#ff6b00", "Low": "#8e909c"}.get(conf, "#8e909c")

        photo_el = (
            html.Img(src=photo, alt=name, style={
                "width": "44px", "height": "44px", "borderRadius": "50%",
                "objectFit": "cover", "border": f"2px solid {conf_color}",
                "flexShrink": "0",
            })
            if photo else
            html.Div(name[:2].upper(), style={
                "width": "44px", "height": "44px", "borderRadius": "50%",
                "display": "flex", "alignItems": "center", "justifyContent": "center",
                "fontSize": "14px", "fontWeight": "900", "color": "#ff6b00",
                "background": "rgba(255,107,0,.15)", "flexShrink": "0",
                "border": f"2px solid {conf_color}",
            })
        )

        card = html.Div([
            # Left: rank badge + photo
            html.Div([
                html.Span(f"#{rank_num}", style={
                    "fontSize": "10px", "fontWeight": "800", "color": "#64748b",
                    "display": "block", "textAlign": "center", "marginBottom": "4px",
                }),
                photo_el,
            ], style={"display": "flex", "flexDirection": "column", "alignItems": "center"}),

            # Center: name + matchup + reasons
            html.Div([
                html.Div([
                    dcc.Link(
                        name,
                        href=f"/player?name={quote_plus(name)}&pitcher={quote_plus(pitcher)}",
                        style={
                            "fontSize": "13px", "fontWeight": "700", "color": "#f1f5f9",
                            "textDecoration": "none", "cursor": "pointer",
                        },
                    ),
                    html.Span(f" · {team}", style={
                        "fontSize": "11px", "color": "#64748b",
                    }),
                ]),
                html.Div(
                    f"vs {opp}" if opp else "",
                    style={"fontSize": "11px", "color": "#94a3b8", "marginTop": "1px"},
                ),
                html.Div([
                    html.Span("▸ " + r, style={"fontSize": "11px", "color": "#94a3b8"})
                    for r in reasons[:2]
                ], style={"marginTop": "4px", "display": "flex", "flexDirection": "column", "gap": "1px"}),
            ], style={"flex": "1", "minWidth": "0", "padding": "0 10px"}),

            # Right: probability
            html.Div([
                html.Div(f"{prob:.1%}", style={
                    "fontSize": "16px", "fontWeight": "800",
                    "color": conf_color, "textAlign": "right",
                }),
                html.Span(conf, style={
                    "fontSize": "9px", "fontWeight": "700",
                    "color": conf_color, "textTransform": "uppercase",
                    "letterSpacing": "1px", "display": "block", "textAlign": "right",
                }),
            ], style={"flexShrink": "0"}),
        ], style={
            "display": "flex", "alignItems": "center",
            "padding": "12px 16px",
            "background": "rgba(255,255,255,.03)",
            "borderRadius": "10px",
            "border": "1px solid rgba(255,255,255,.06)",
            "gap": "8px",
        })
        cards.append(card)

    if not cards:
        return []

    return html.Div([
        html.Div([
            html.H4("Featured Players", style={
                "margin": 0, "fontSize": "15px", "fontWeight": "800",
                "color": "var(--primary)",
            }),
            html.P("Next 10 ranked picks with model reasoning",
                   style={"margin": "2px 0 0", "fontSize": "12px",
                          "color": "var(--on-surface-variant)"}),
        ], className="table-title-bar"),
        html.Div(cards, style={
            "display": "grid",
            "gridTemplateColumns": "repeat(auto-fill, minmax(340px, 1fr))",
            "gap": "8px",
            "padding": "0 0 8px",
        }),
    ], className="table-container")


def _build_hero_spotlight(df: pd.DataFrame):
    if df.empty:
        return _empty_hero()

    row  = df.iloc[0]
    name = str(row.get("Player", ""))
    park = float(row.get("Park_Factor", 100) or 100)
    wind = str(row.get("Wind_Direction", "calm") or "calm")

    try:
        from utils.player_photos import batch_headshot_urls
        photo = batch_headshot_urls([name], width=120).get(name, "")
    except Exception:
        photo = ""

    stats = _lookup_batting_stats(name)

    # ── Insight text: SHAP explanation → fallback to heuristic ───────────────
    shap_insight = str(row.get("Insight_Text", "") or "")
    if shap_insight and shap_insight != "Model-selected pick":
        context = shap_insight.capitalize() + "."
    else:
        # Heuristic fallback when no SHAP data available
        parts = []
        ev = stats.get("ev")
        if ev and float(ev) > 0:
            parts.append(f"Averaging {float(ev):.1f} mph exit velocity")
        barrel = stats.get("barrel_pct")
        if barrel and float(barrel) >= 0.10:
            parts.append(f"{float(barrel)*100:.0f}% barrel rate")
        if wind.lower() == "out":
            parts.append("wind blowing out to the outfield")
        elif wind.lower() == "in":
            parts.append("headwind in play")
        if park >= 108:
            parts.append("hitter-friendly venue")
        elif park <= 92:
            parts.append("pitcher-friendly park")
        context = ". ".join(parts[:2]).capitalize() + "." if parts else "Model's highest HR probability pick today."

    # ── Stat chips: AVG | HR (featured) | SLG ────────────────────────────────
    avg_val = f".{int(round(float(stats['avg']), 3) * 1000):03d}" if stats.get("avg") else "—"
    hr_val  = str(stats.get("hr", "—"))
    slg_val = f".{int(round(float(stats['slg']), 3) * 1000):03d}" if stats.get("slg") else "—"

    chip_els = [
        html.Div([
            html.Span("AVG", className="hero-stat-label"),
            html.Div(avg_val, className="hero-stat-value"),
        ], className="hero-stat-chip"),
        html.Div([
            html.Span("HR", className="hero-stat-label"),
            html.Div(hr_val, className="hero-stat-value"),
        ], className="hero-stat-chip featured"),
        html.Div([
            html.Span("SLG", className="hero-stat-label"),
            html.Div(slg_val, className="hero-stat-value"),
        ], className="hero-stat-chip"),
    ]

    return html.Div([
        html.Img(src=photo, alt=name, className="hero-spotlight-bg") if photo else None,
        html.Div(className="hero-spotlight-gradient"),
        html.Div([
            html.Div([
                html.Span("Player to Watch", className="hero-badge"),
                html.H3(name, className="hero-player-name"),
                html.P(context, className="hero-player-desc"),
            ]),
            html.Div(chip_els, className="hero-stat-chips"),
        ], className="hero-spotlight-content"),
    ], className="hero-spotlight")


def _empty_hero():
    return html.Div([
        html.Div(className="hero-spotlight-gradient"),
        html.Div([
            html.Div([
                html.Span("Player to Watch", className="hero-badge"),
                html.H3("No Data Yet", className="hero-player-name"),
                html.P("Run the morning scheduler to load today's predictions.",
                       className="hero-player-desc"),
            ]),
        ], className="hero-spotlight-content"),
    ], className="hero-spotlight")


_PARK_NAMES = {
    "ARI": "Chase Field",       "ATL": "Truist Park",
    "BAL": "Camden Yards",      "BOS": "Fenway Park",
    "CHC": "Wrigley Field",     "CHW": "Guaranteed Rate",
    "CIN": "Great American",    "CLE": "Progressive Field",
    "COL": "Coors Field",       "DET": "Comerica Park",
    "HOU": "Minute Maid Park",  "KC":  "Kauffman Stadium",
    "LAA": "Angel Stadium",     "LAD": "Dodger Stadium",
    "MIA": "LoanDepot Park",    "MIL": "American Family",
    "MIN": "Target Field",      "NYM": "Citi Field",
    "NYY": "Yankee Stadium",    "OAK": "Sutter Health Park",
    "PHI": "Citizens Bank Park","PIT": "PNC Park",
    "SD":  "Petco Park",        "SF":  "Oracle Park",
    "SEA": "T-Mobile Park",     "STL": "Busch Stadium",
    "TB":  "Tropicana Field",   "TEX": "Globe Life Field",
    "TOR": "Rogers Centre",     "WSH": "Nationals Park",
}


# ─── Venue analysis card ──────────────────────────────────────────────────────

def _build_venue_card(df: pd.DataFrame):
    if df.empty:
        return _empty_venue()

    # Find the game with highest park factor (most interesting venue)
    pf_col = pd.to_numeric(df["Park_Factor"], errors="coerce")
    idx = pf_col.idxmax() if pf_col.notna().any() else df.index[0]
    row = df.loc[idx]

    park  = float(row.get("Park_Factor", 100) or 100)
    home_team = str(row.get("Team", "") if row.get("Home_Game") in (True, 1) else row.get("Opponent", ""))
    wind  = str(row.get("Wind_Direction", "calm") or "calm")
    speed = row.get("Wind_Speed_MPH", 0)
    try:
        speed = int(float(speed or 0))
    except (TypeError, ValueError):
        speed = 0

    park_name  = _PARK_NAMES.get(home_team, f"{home_team} Ballpark")
    park_delta = park - 100
    delta_sign = "+" if park_delta >= 0 else ""

    # Body text like the reference design
    if park_delta >= 10:
        body_text = (
            f"{park_name} currently showing {delta_sign}{park_delta:.0f}% HR volatility "
            f"based on park dimensions and atmospheric conditions."
        )
    elif park_delta >= 5:
        body_text = (
            f"{park_name} slightly favors hitters — park factor {int(park)} "
            f"({delta_sign}{park_delta:.0f}% above league average)."
        )
    elif park_delta <= -5:
        body_text = (
            f"{park_name} suppresses home runs — park factor {int(park)} "
            f"({park_delta:.0f}% below league average)."
        )
    else:
        body_text = (
            f"{park_name} — neutral park factor {int(park)}. "
            f"Context factors in play today: wind {wind} at {speed} mph."
        )

    # Wind bar display
    wind_lower = wind.lower()
    if wind_lower == "out":
        bar_pct    = min(speed * 5, 100)
        wind_color = "var(--green)"
        wind_label = f"OUT TO CENTER"
    elif wind_lower == "in":
        bar_pct    = min(speed * 5, 100)
        wind_color = "var(--red)"
        wind_label = f"IN FROM CENTER"
    elif wind_lower == "cross":
        bar_pct    = min(speed * 3, 70)
        wind_color = "var(--primary)"
        wind_label = f"CROSSWIND"
    else:
        bar_pct    = min(speed * 2, 30)
        wind_color = "var(--on-surface-variant)"
        wind_label = f"CALM"

    return html.Div([
        html.P("Venue Analysis", className="venue-card-title"),
        html.P(body_text, className="venue-card-body"),
        html.Div([
            html.Div([
                html.Span("Wind Direction", className="venue-stat-key"),
                html.Span(wind_label, className="venue-stat-val",
                          style={"color": wind_color}),
            ], className="venue-stat-header"),
            html.Div([
                html.Div(style={
                    "width": f"{bar_pct}%", "height": "100%",
                    "background": wind_color, "borderRadius": "2px",
                    "transition": "width .6s",
                }),
            ], style={
                "height": "6px", "background": "var(--surface)",
                "borderRadius": "3px", "overflow": "hidden",
            }),
        ], className="venue-stat-row"),
    ], className="venue-card")


def _empty_venue():
    return html.Div([
        html.P("Venue Analysis", className="venue-card-title"),
        html.P("No game data available yet.", className="venue-card-body"),
    ], className="venue-card")


# ─── Daily 3-leg parlay ───────────────────────────────────────────────────────

def _build_parlay_card(df: pd.DataFrame):
    """Build the full parlays section: 5 three-leg + 5 two-leg parlays."""
    if df.empty:
        return []

    import itertools
    import math

    pool = df.sort_values("HR_Probability", ascending=False).head(12)
    if len(pool) < 2:
        return []

    rows = [row for _, row in pool.iterrows()]

    def _combined_prob(legs):
        return math.prod(float(r["HR_Probability"]) for r in legs)

    def _parlay_mini_card(legs, parlay_num: int, accent: str):
        combined = _combined_prob(legs)
        leg_items = []
        for i, row in enumerate(legs, 1):
            name = str(row.get("Player", ""))
            team = str(row.get("Team", ""))
            opp  = str(row.get("Opponent", ""))
            prob = float(row.get("HR_Probability", 0))
            conf = str(row.get("Confidence", ""))
            badge_color = "#22c55e" if conf == "High" else "#ff6b00"
            leg_items.append(html.Div([
                html.Div(str(i), style={
                    "width": "18px", "height": "18px", "borderRadius": "50%",
                    "background": "rgba(255,107,0,.12)",
                    "border": f"1px solid {accent}",
                    "color": accent, "fontSize": "10px", "fontWeight": "700",
                    "display": "flex", "alignItems": "center", "justifyContent": "center",
                    "flexShrink": "0",
                }),
                html.Div([
                    html.Span(name, style={"fontSize": "12px", "fontWeight": "600", "color": "#f0dfd8"}),
                    html.Span(f" {team} vs {opp}", style={
                        "fontSize": "10px", "color": "var(--on-surface-variant)", "marginLeft": "4px",
                    }),
                ], style={"flex": "1", "minWidth": "0", "overflow": "hidden"}),
                html.Div([
                    html.Span(f"{prob:.1%}", style={"fontSize": "11px", "fontWeight": "700", "color": accent}),
                    html.Span(conf, style={
                        "fontSize": "9px", "color": badge_color,
                        "marginLeft": "5px", "border": f"1px solid {badge_color}",
                        "borderRadius": "3px", "padding": "1px 4px",
                    }),
                ], style={"display": "flex", "alignItems": "center", "flexShrink": "0"}),
            ], style={
                "display": "flex", "alignItems": "center", "gap": "8px",
                "padding": "5px 0", "borderBottom": "1px solid rgba(255,255,255,.04)",
            }))

        return html.Div([
            html.Div([
                html.Span(f"Parlay {parlay_num}", style={
                    "fontSize": "10px", "fontWeight": "800", "color": accent,
                    "textTransform": "uppercase", "letterSpacing": "1px",
                }),
                html.Div([
                    html.Span("Combined: ", style={"fontSize": "10px", "color": "var(--on-surface-variant)"}),
                    html.Span(f"{combined:.2%}", style={"fontSize": "13px", "fontWeight": "800", "color": accent}),
                ], style={"display": "flex", "alignItems": "center", "gap": "4px"}),
            ], style={
                "display": "flex", "justifyContent": "space-between", "alignItems": "center",
                "marginBottom": "6px",
            }),
            html.Div(leg_items),
        ], style={
            "background": "var(--surface-container)",
            "border": f"1px solid {accent}33",
            "borderRadius": "10px", "padding": "12px 14px",
        })

    # ── 3-leg parlays ──────────────────────────────────────────────────────────
    three_leg_section: list = []
    if len(rows) >= 3:
        three_combos = sorted(
            itertools.combinations(rows, 3),
            key=_combined_prob,
            reverse=True,
        )[:5]
        three_cards = [
            _parlay_mini_card(list(legs), i + 1, "#ff6b00")
            for i, legs in enumerate(three_combos)
        ]
        three_leg_section = [
            html.Div([
                html.H4("3-Leg HR Parlays", style={
                    "margin": "0 0 4px", "fontSize": "14px", "fontWeight": "800",
                    "color": "#ff6b00",
                }),
                html.P("Top 5 three-player HR combos by combined probability",
                       style={"margin": 0, "fontSize": "11px", "color": "var(--on-surface-variant)"}),
            ], style={"marginBottom": "14px"}),
            html.Div(three_cards, style={
                "display": "grid",
                "gridTemplateColumns": "repeat(auto-fill, minmax(320px, 1fr))",
                "gap": "12px",
            }),
        ]

    # ── 2-leg parlays ──────────────────────────────────────────────────────────
    two_leg_section: list = []
    if len(rows) >= 2:
        two_combos = sorted(
            itertools.combinations(rows, 2),
            key=_combined_prob,
            reverse=True,
        )[:5]
        two_cards = [
            _parlay_mini_card(list(legs), i + 1, "#22c55e")
            for i, legs in enumerate(two_combos)
        ]
        two_leg_section = [
            html.Div([
                html.H4("2-Leg HR Parlays", style={
                    "margin": "0 0 4px", "fontSize": "14px", "fontWeight": "800",
                    "color": "#22c55e",
                }),
                html.P("Top 5 two-player HR combos by combined probability",
                       style={"margin": 0, "fontSize": "11px", "color": "var(--on-surface-variant)"}),
            ], style={"marginBottom": "14px"}),
            html.Div(two_cards, style={
                "display": "grid",
                "gridTemplateColumns": "repeat(auto-fill, minmax(320px, 1fr))",
                "gap": "12px",
            }),
        ]

    if not three_leg_section and not two_leg_section:
        return []

    return html.Div([
        # Section header
        html.Div([
            html.Div([
                html.H4("Today's HR Parlays", style={
                    "margin": 0, "fontSize": "15px", "fontWeight": "800",
                }),
                html.P("Model-generated parlay suggestions — ranked by combined HR probability",
                       style={"margin": "2px 0 0", "fontSize": "12px",
                              "color": "var(--on-surface-variant)"}),
            ]),
        ], className="table-title-bar"),

        html.Div([
            *three_leg_section,
            html.Div(style={"height": "20px"}) if three_leg_section and two_leg_section else None,
            *two_leg_section,
        ], style={"padding": "16px"}),

        html.Div(
            "Probabilities are independent estimates. For entertainment only.",
            style={"fontSize": "10px", "color": "rgba(255,255,255,.2)", "padding": "8px 16px"},
        ),
    ], className="table-container", style={
        "border": "1px solid rgba(255,107,0,.2)",
        "background": "linear-gradient(135deg, rgba(255,107,0,.05) 0%, var(--surface-container) 60%)",
    })


# ─── Weather banner ───────────────────────────────────────────────────────────

def _build_weather_banner(df: pd.DataFrame):
    if df.empty:
        return []

    wind_counts = df["Wind_Direction"].value_counts()
    dominant_wind = wind_counts.index[0] if not wind_counts.empty else "calm"
    avg_temp = df["Temp_F"].apply(pd.to_numeric, errors="coerce").mean()
    avg_park = df["Park_Factor"].apply(pd.to_numeric, errors="coerce").mean()

    lhb_factor = round(float(avg_park or 100), 0)
    rhb_factor = round(float(avg_park or 100) * 0.98, 0)

    wind_label = dominant_wind.title()
    wind_cls   = "boost" if dominant_wind == "out" else ("penalty" if dominant_wind == "in" else "normal")
    temp_cls   = "boost" if (avg_temp or 72) >= 80 else ("penalty" if (avg_temp or 72) <= 45 else "normal")

    return html.Div([
        html.Span("Weather Factor", className="weather-banner-label"),
        html.Span([
            html.Span("LHB ", style={"fontSize": "10px", "opacity": ".7"}),
            f"{int(lhb_factor)}%",
        ], className=f"weather-factor-chip {'boost' if lhb_factor > 100 else 'normal'}"),
        html.Span([
            html.Span("RHB ", style={"fontSize": "10px", "opacity": ".7"}),
            f"{int(rhb_factor)}%",
        ], className=f"weather-factor-chip {'boost' if rhb_factor > 100 else 'normal'}"),
        html.Span("·", style={"color": "rgba(255,255,255,.2)"}),
        html.Span([
            html.Span("Wind ", style={"fontSize": "10px", "opacity": ".7"}),
            wind_label,
        ], className=f"weather-factor-chip {wind_cls}"),
        html.Span([
            html.Span("Temp ", style={"fontSize": "10px", "opacity": ".7"}),
            f"{int(avg_temp or 72)}°F",
        ], className=f"weather-factor-chip {temp_cls}"),
    ], className="weather-banner")


# ─── Empty states ─────────────────────────────────────────────────────────────

def _empty_picks():
    return html.Div([
        html.Div([
            html.Div(style={
                "fontSize": "32px", "marginBottom": "8px",
                "color": "var(--on-surface-variant)",
            }),
            html.Div("No predictions yet for today",
                     style={"fontWeight": "700", "fontSize": "14px",
                            "color": "var(--on-surface)"}),
            html.Div("The scheduler runs at 9:00 AM",
                     style={"fontSize": "12px", "color": "var(--on-surface-variant)",
                            "marginTop": "4px"}),
        ], style={"textAlign": "center", "padding": "48px 24px"}),
    ])


# ─── Summary cards ────────────────────────────────────────────────────────────

def _build_cards(df: pd.DataFrame):
    total = len(df)
    high  = int((df["Confidence"] == "High").sum())
    med   = int((df["Confidence"] == "Medium").sum())
    avg_p = float(df["HR_Probability"].mean()) if total > 0 else 0

    def card(title, value, cls=""):
        return html.Div([
            html.Div(str(value), className="summary-card-value",
                     style={"color": "#f0dfd8" if not cls else
                            "#22c55e" if cls == "green" else
                            "#ff6b00" if cls == "orange" else "#ff6b00"}),
            html.Div(title, className="summary-card-label"),
        ], className=f"summary-card {cls}")

    return [
        card("Total Players", total),
        card("High Confidence", high, "green"),
        card("Medium Confidence", med, "orange"),
        card("Avg HR Probability", f"{avg_p:.1%}", "amber"),
    ]


def _empty_cards():
    return [html.Div(
        "No predictions yet for today — scheduler runs at 9:00 AM.",
        style={"color": "var(--on-surface-variant)", "padding": "16px",
               "fontStyle": "italic", "fontSize": "14px"},
    )]


# ─── Distribution chart ───────────────────────────────────────────────────────

def _build_dist_chart(df: pd.DataFrame):
    if df.empty:
        return _empty_chart()
    fig = px.histogram(
        df, x="HR_Probability", color="Confidence",
        color_discrete_map=CONF_COLORS,
        nbins=20,
        labels={"HR_Probability": "HR Probability", "count": "Players"},
    )
    fig.update_layout(
        paper_bgcolor="#131313", plot_bgcolor="#1a1a1a",
        legend_title_text="Confidence",
        margin=dict(l=40, r=20, t=10, b=40),
        height=260, bargap=0.08,
        font={"family": "Manrope, sans-serif", "size": 12, "color": "#f0dfd8"},
        xaxis={"gridcolor": "rgba(255,255,255,.05)", "color": "#adaaaa"},
        yaxis={"gridcolor": "rgba(255,255,255,.05)", "color": "#adaaaa"},
    )
    fig.update_xaxes(tickformat=".0%")
    return fig


def _empty_chart():
    fig = go.Figure()
    fig.add_annotation(
        text="No predictions loaded yet",
        xref="paper", yref="paper", x=0.5, y=0.5, showarrow=False,
        font={"size": 14, "color": "#6b6b6b"},
    )
    fig.update_layout(height=260, paper_bgcolor="#131313", plot_bgcolor="#1a1a1a",
                      margin=dict(l=40, r=20, t=10, b=40))
    return fig


# ─── Player modal helpers ─────────────────────────────────────────────────────

def _lookup_pitcher_stats(pitcher_name: str) -> dict:
    """Return FIP, xFIP, HR/9, HR/FB, GB%, K/9 for a pitcher from FanGraphs 2025."""
    if not pitcher_name:
        return {}
    try:
        from utils.data_fetcher import fetch_fangraphs_pitching
        df = fetch_fangraphs_pitching(2025)
        if df is None or df.empty:
            return {}
        parts = pitcher_name.split()
        if len(parts) >= 2:
            mask = (
                df["Name"].str.contains(parts[0], case=False, na=False) &
                df["Name"].str.contains(parts[-1], case=False, na=False)
            )
        else:
            mask = df["Name"].str.contains(pitcher_name, case=False, na=False)
        row = df[mask]
        if row.empty:
            return {}
        r = row.iloc[0]
        return {
            "era":   r.get("ERA"),
            "fip":   r.get("FIP"),
            "xfip":  r.get("xFIP"),
            "hr9":   r.get("HR/9"),
            "hr_fb": r.get("HR/FB"),
            "gb_pct": r.get("GB%"),
            "k9":    r.get("K/9"),
            "bb9":   r.get("BB/9"),
        }
    except Exception:
        return {}


def _stat_chip(label: str, value: str, highlight: bool = False):
    return html.Div([
        html.Div(value, style={
            "fontSize": "18px", "fontWeight": "800",
            "color": "#ff6b00" if highlight else "#f0dfd8",
            "fontFamily": "Manrope, sans-serif",
        }),
        html.Div(label, style={
            "fontSize": "10px", "color": "#8e909c",
            "textTransform": "uppercase", "letterSpacing": "0.8px",
            "marginTop": "2px",
        }),
    ], className="modal-stat-chip")


def _build_player_modal(row: dict) -> list:
    """Build modal content from a predictions table row."""
    player_name  = str(row.get("Player", ""))
    pitcher_name = str(row.get("Pitcher", ""))
    team         = str(row.get("Team", ""))
    opponent     = str(row.get("Opponent", ""))
    hr_prob      = float(row.get("HR_Probability", 0) or 0)
    park_factor  = float(row.get("Park_Factor", 100) or 100)
    temp_f       = row.get("Temp_F")
    wind_dir     = str(row.get("Wind_Direction", "calm") or "calm")
    wind_mph     = row.get("Wind_Speed_MPH", 0)
    confidence   = str(row.get("Confidence", ""))
    is_indoor    = row.get("Is_Indoor")

    conf_color = {"High": "#22c55e", "Medium": "#ff6b00", "Low": "#8e909c"}.get(confidence, "#8e909c")

    # Photo
    try:
        from utils.player_photos import batch_headshot_urls
        photo = batch_headshot_urls([player_name], width=100).get(player_name, "")
    except Exception:
        photo = ""

    # Season stats
    batting = _lookup_batting_stats(player_name)

    # Pitcher stats
    pitcher = _lookup_pitcher_stats(pitcher_name)

    # Park name
    home_team = team if row.get("Home_Game") in (True, 1) else opponent
    park_name = _PARK_NAMES.get(home_team, f"{home_team} Ballpark")

    # ── Header ───────────────────────────────────────────────────────────────
    header = html.Div([
        html.Img(src=photo, alt=player_name, className="modal-player-photo") if photo
        else html.Div(player_name[:2].upper(), className="modal-player-photo modal-player-initials"),
        html.Div([
            html.H3(player_name, className="modal-player-name"),
            html.Div(f"{team}  ·  vs {opponent}", style={
                "fontSize": "13px", "color": "#8e909c", "marginBottom": "6px",
            }),
            html.Div([
                html.Span(f"{hr_prob:.1%} HR Prob", style={
                    "color": "#ff6b00", "fontWeight": "800", "fontSize": "15px",
                    "marginRight": "10px",
                }),
                html.Span(confidence, style={
                    "color": conf_color, "fontWeight": "700", "fontSize": "12px",
                    "border": f"1px solid {conf_color}", "borderRadius": "4px",
                    "padding": "2px 8px",
                }),
            ]),
        ]),
    ], className="modal-header-row")

    # ── Batting stat chips ────────────────────────────────────────────────────
    avg_s   = f".{int(round(float(batting['avg']), 3) * 1000):03d}" if batting.get("avg") else "—"
    slg_s   = f".{int(round(float(batting['slg']), 3) * 1000):03d}" if batting.get("slg") else "—"
    ev_s    = f"{float(batting['ev']):.1f}" if batting.get("ev") else "—"
    bar_s   = f"{float(batting['barrel_pct'])*100:.1f}%" if batting.get("barrel_pct") else "—"
    hr_fb_s = f"{float(batting['hr_fb'])*100:.1f}%" if batting.get("hr_fb") else "—"

    batting_chips = html.Div([
        _stat_chip("HR (2025)", str(batting.get("hr", "—")), highlight=True),
        _stat_chip("AVG", avg_s),
        _stat_chip("SLG", slg_s),
        _stat_chip("Exit Vel", ev_s + " mph" if batting.get("ev") else "—"),
        _stat_chip("Barrel%", bar_s),
        _stat_chip("HR/FB", hr_fb_s),
    ], className="modal-stat-chips")

    # ── Pitcher stat chips ────────────────────────────────────────────────────
    if pitcher:
        era_s  = f"{float(pitcher['era']):.2f}" if pitcher.get("era") else "—"
        fip_s  = f"{float(pitcher['fip']):.2f}" if pitcher.get("fip") else "—"
        hr9_s  = f"{float(pitcher['hr9']):.2f}" if pitcher.get("hr9") else "—"
        hr_fb_p = f"{float(pitcher['hr_fb'])*100:.1f}%" if pitcher.get("hr_fb") else "—"
        gb_s   = f"{float(pitcher['gb_pct'])*100:.1f}%" if pitcher.get("gb_pct") else "—"
        k9_s   = f"{float(pitcher['k9']):.1f}" if pitcher.get("k9") else "—"
        pitcher_chips = html.Div([
            _stat_chip("ERA", era_s),
            _stat_chip("FIP", fip_s),
            _stat_chip("HR/9", hr9_s, highlight=True),
            _stat_chip("HR/FB", hr_fb_p),
            _stat_chip("GB%", gb_s),
            _stat_chip("K/9", k9_s),
        ], className="modal-stat-chips")
    else:
        pitcher_chips = html.Div("No 2025 data found for this pitcher",
                                  style={"color": "#8e909c", "fontSize": "13px", "padding": "8px 0"})

    # ── Park stat chips ───────────────────────────────────────────────────────
    pf_delta = park_factor - 100
    pf_color = "#22c55e" if pf_delta >= 5 else "#ef4444" if pf_delta <= -5 else "#ff6b00"
    pf_s = f"{int(park_factor)} ({'+' if pf_delta >= 0 else ''}{pf_delta:.0f}%)"

    wind_color = {"out": "#22c55e", "in": "#ef4444", "cross": "#ff6b00"}.get(wind_dir.lower(), "#8e909c")
    temp_s = f"{int(float(temp_f))}°F" if temp_f else ("Indoor" if is_indoor else "—")
    wind_s = f"{int(float(wind_mph or 0))} mph {wind_dir.title()}"

    park_chips = html.Div([
        html.Div([
            html.Div(pf_s, style={
                "fontSize": "18px", "fontWeight": "800",
                "color": pf_color, "fontFamily": "Manrope, sans-serif",
            }),
            html.Div("Park Factor", style={
                "fontSize": "10px", "color": "#8e909c",
                "textTransform": "uppercase", "letterSpacing": "0.8px", "marginTop": "2px",
            }),
        ], className="modal-stat-chip"),
        _stat_chip("Temp", temp_s),
        html.Div([
            html.Div(wind_s, style={
                "fontSize": "15px", "fontWeight": "800",
                "color": wind_color, "fontFamily": "Manrope, sans-serif",
            }),
            html.Div("Wind", style={
                "fontSize": "10px", "color": "#8e909c",
                "textTransform": "uppercase", "letterSpacing": "0.8px", "marginTop": "2px",
            }),
        ], className="modal-stat-chip"),
        _stat_chip("Park", park_name[:18] + ("…" if len(park_name) > 18 else "")),
    ], className="modal-stat-chips")

    def section(title: str, content):
        return html.Div([
            html.H4(title, className="modal-section-title"),
            content,
        ], className="modal-section")

    return [
        header,
        html.Hr(className="modal-divider"),
        section("Season Stats (2025)", batting_chips),
        section(f"Pitcher: {pitcher_name}", pitcher_chips),
        section(f"At {park_name}", park_chips),
    ]


# ─── Player modal callback ────────────────────────────────────────────────────

@callback(
    Output("player-modal-overlay", "style"),
    Output("player-modal-content", "children"),
    Input("predictions-table", "active_cell"),
    Input("modal-close-btn", "n_clicks"),
    State("predictions-table", "data"),
    prevent_initial_call=True,
)
def handle_player_modal(active_cell, _close, data):
    if ctx.triggered_id == "modal-close-btn":
        return {"display": "none"}, []

    if not active_cell or active_cell.get("column_id") != "Player":
        return {"display": "none"}, []

    if not data:
        return {"display": "none"}, []

    row_idx = active_cell.get("row", 0)
    if row_idx >= len(data):
        return {"display": "none"}, []

    row = data[row_idx]
    content = _build_player_modal(row)
    return {"display": "flex"}, content

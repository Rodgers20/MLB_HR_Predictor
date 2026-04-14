"""
MLB HR Predictor — Dash Dashboard
Entry point: python dashboard/app.py
Navigate to http://localhost:8050
"""

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from dotenv import load_dotenv
load_dotenv()

import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

import dash
from dash import Input, Output, dcc, html

from dashboard.pages import history, model_perf, player, player_stats, today

EXCEL_PATH = Path("MLB_HR_Predictions.xlsx")
SEASON_YEAR = 2026


def _load_hr_leaderboard() -> list[tuple[str, int]]:
    """Return list of (player, season_hr_count) sorted desc, only players with ≥1 HR.

    Source: FanGraphs season batting stats (fetched daily, authoritative HR totals).
    Falls back to Excel tracker if FanGraphs CSV is unavailable.
    """
    fg_path = Path("data/raw/fangraphs_batting_2026.csv")
    if fg_path.exists():
        try:
            df = pd.read_csv(fg_path, usecols=["Name", "Team", "HR"])
            df["HR"] = pd.to_numeric(df["HR"], errors="coerce").fillna(0).astype(int)
            leaders = (
                df[df["HR"] > 0]
                .sort_values("HR", ascending=False)
                .drop_duplicates(subset=["Name"])
            )
            return list(zip(leaders["Name"], leaders["HR"]))
        except Exception:
            pass  # fall through to Excel fallback

    # Fallback: read from Excel Predictions sheet
    if not EXCEL_PATH.exists():
        return []
    try:
        wb   = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        ws   = wb["Predictions"]
        data = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
    except Exception:
        return []
    if not data:
        return []
    cols = [
        "Date", "Player", "Team", "Opponent", "Pitcher",
        "HR_Probability", "Confidence", "Park_Factor",
        "Temp_F", "Wind_Speed_MPH", "Wind_Direction", "Is_Indoor",
        "Home_Game", "Predicted_HRs", "Actual_HRs", "Hit",
        "Insight_Text",
    ]
    row_len = len(data[0]) if data else 0
    df = pd.DataFrame(data, columns=cols[:row_len] if row_len <= len(cols) else cols)
    df["Date"]       = pd.to_datetime(df["Date"], errors="coerce")
    df["Actual_HRs"] = pd.to_numeric(df["Actual_HRs"], errors="coerce")
    season = df[(df["Date"].dt.year == SEASON_YEAR) & (df["Actual_HRs"] > 0)]
    if season.empty:
        return []
    season = (
        season
        .sort_values("Actual_HRs", ascending=False)
        .drop_duplicates(subset=["Player", "Date"], keep="first")
    )
    leaderboard = (
        season.groupby("Player")["Actual_HRs"]
        .sum()
        .astype(int)
        .sort_values(ascending=False)
    )
    return list(leaderboard.items())

BOOTSTRAP_CDN = "https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css"

app = dash.Dash(
    __name__,
    external_stylesheets=[BOOTSTRAP_CDN],
    suppress_callback_exceptions=True,
    title="MLB HR Predictor",
)
server = app.server

PAGES = [
    {"href": "/",              "label": "Dashboard",         "icon": "grid_view"},
    {"href": "/player",        "label": "Predictions",       "icon": "insights"},
    {"href": "/performance",   "label": "Model Performance", "icon": "emoji_events"},
    {"href": "/player-stats",  "label": "Player Tracker",    "icon": "person_search"},
    {"href": "/history",       "label": "History",           "icon": "history"},
]


def _nav_links(pathname: str):
    return html.Div([
        dcc.Link(
            [
                html.Span(p["icon"], className="material-symbols-outlined"),
                html.Span(p["label"], style={"fontSize": "13.5px"}),
            ],
            href=p["href"],
            className="nav-link active" if pathname == p["href"] else "nav-link",
            style={"gap": "12px"},
        )
        for p in PAGES
    ])


def _hr_leaderboard_widget(leaderboard: list[tuple[str, int]]):
    """Sidebar HR leaderboard widget."""
    header = html.Div("2026 HR Leaders", style={
        "fontFamily": "'Inter', sans-serif",
        "fontSize": "10px", "fontWeight": "700",
        "color": "rgba(255,255,255,.3)",
        "textTransform": "uppercase", "letterSpacing": "1.2px",
        "marginBottom": "10px",
    })

    if not leaderboard:
        body = html.Div("No confirmed HRs yet", style={
            "fontSize": "11px", "color": "rgba(255,255,255,.25)",
            "fontStyle": "italic",
        })
    else:
        rows = []
        for rank, (player, hrs) in enumerate(leaderboard, 1):
            parts = player.split()
            short = f"{parts[0][0]}. {' '.join(parts[1:])}" if len(parts) > 1 else player
            bar_pct = min(100, hrs * 12)  # visual bar, max ~8 HRs
            rows.append(html.Div([
                html.Div([
                    html.Span(f"{rank}", style={
                        "fontFamily": "'Manrope', sans-serif",
                        "fontSize": "10px", "fontWeight": "700",
                        "color": "rgba(255,255,255,.2)",
                        "width": "14px", "flexShrink": "0",
                    }),
                    html.Span(short, style={
                        "fontFamily": "'Inter', sans-serif",
                        "fontSize": "12px", "color": "#f0f4f8",
                        "flex": "1", "overflow": "hidden",
                        "textOverflow": "ellipsis", "whiteSpace": "nowrap",
                    }),
                    html.Span(str(hrs), style={
                        "fontFamily": "'Manrope', sans-serif",
                        "fontSize": "13px", "fontWeight": "800",
                        "color": "#ff6b00", "flexShrink": "0",
                    }),
                ], style={"display": "flex", "alignItems": "center", "gap": "6px", "marginBottom": "3px"}),
                # Thin progress bar
                html.Div(html.Div(style={
                    "height": "100%", "width": f"{bar_pct}%",
                    "background": "linear-gradient(90deg, #4d2000, #ff6b00)",
                    "borderRadius": "2px",
                }), style={
                    "height": "2px", "background": "rgba(255,255,255,.06)",
                    "borderRadius": "2px", "overflow": "hidden", "marginBottom": "6px",
                }),
            ]))
        body = html.Div(rows, style={"maxHeight": "200px", "overflowY": "auto"})

    return html.Div([header, body], style={
        "margin": "16px 12px 0",
        "padding": "14px 16px",
        "background": "rgba(255,107,0,.05)",
        "border": "1px solid rgba(255,107,0,.15)",
        "borderRadius": "10px",
    })


def sidebar(pathname: str = "/", leaderboard: list | None = None):
    return [
        html.Div([
            html.P("Editorial Elite", className="sidebar-brand-title"),
            html.P("Sabermetric Gallery", className="sidebar-brand-sub"),
            html.Div("2026 Season", className="sidebar-season-badge"),
        ], className="sidebar-header"),

        html.Div("Navigation", className="nav-section-label"),
        _nav_links(pathname),

        _hr_leaderboard_widget(leaderboard or []),

        html.Div(style={"flex": "1"}),

        # New Analysis CTA button
        html.Div([
            html.Button("New Analysis", style={
                "width": "100%", "padding": "12px",
                "background": "linear-gradient(135deg, #ff7524, #ff9159)",
                "border": "none", "borderRadius": "8px",
                "color": "#531e00", "fontSize": "13px", "fontWeight": "700",
                "cursor": "pointer", "letterSpacing": ".3px",
            }),
        ], style={"padding": "0 12px 12px"}),

        html.Div([
            html.Div("XGBoost · Statcast · FanGraphs",
                     style={"fontFamily": "'Manrope', sans-serif",
                            "fontSize": "10px", "color": "rgba(255,255,255,.18)",
                            "padding": "12px 20px",
                            "borderTop": "1px solid rgba(255,255,255,.05)"}),
        ]),
    ]


# ─── App Layout ───────────────────────────────────────────────────────────────

app.layout = html.Div([
    dcc.Location(id="url"),
    dcc.Interval(id="hr-counter-interval", interval=5 * 60 * 1000, n_intervals=0),
    html.Div(sidebar("/"), id="sidebar"),
    html.Div(id="main-content", style={"marginLeft": "240px", "flex": "1"}),
], style={"display": "flex", "minHeight": "100vh"})


# ─── Page Routing ─────────────────────────────────────────────────────────────

@app.callback(
    Output("main-content", "children"),
    Input("url", "pathname"),
    Input("url", "search"),
)
def render_page(pathname, search):
    if pathname == "/player":
        player_name  = None
        pitcher_name = None
        if search:
            from urllib.parse import parse_qs
            params = parse_qs(search.lstrip("?"))
            player_name  = params.get("name",    [None])[0]
            pitcher_name = params.get("pitcher", [None])[0]
        return player.layout(player_name=player_name, pitcher_name=pitcher_name)
    elif pathname == "/performance":
        return model_perf.layout()
    elif pathname == "/player-stats":
        return player_stats.layout()
    elif pathname == "/history":
        return history.layout()
    else:
        return today.layout()


# ─── Active nav highlighting ──────────────────────────────────────────────────

@app.callback(
    Output("sidebar", "children"),
    Input("url", "pathname"),
    Input("hr-counter-interval", "n_intervals"),
)
def update_nav(pathname, _n):
    return sidebar(pathname, leaderboard=_load_hr_leaderboard())


# ─── Run ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    host = os.getenv("DASH_HOST", "0.0.0.0")
    port = int(os.getenv("DASH_PORT", 8050))
    debug = os.getenv("DASH_DEBUG", "False").lower() == "true"
    print(f"\n  MLB HR Predictor → http://localhost:{port}\n")
    app.run(host=host, port=port, debug=debug)

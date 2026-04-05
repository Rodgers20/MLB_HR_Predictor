"""
Prediction tracker — writes daily predictions and results to MLB_HR_Predictions.xlsx.
Sheets: Predictions, Model_Performance, Feature_Importance.
"""

import logging
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)

EXCEL_PATH = Path("MLB_HR_Predictions.xlsx")

PRED_COLUMNS = [
    "Date", "Player", "Team", "Opponent", "Pitcher",
    "HR_Probability", "Confidence", "Park_Factor",
    "Temp_F", "Wind_Speed_MPH", "Wind_Direction", "Is_Indoor",
    "Home_Game", "Predicted_HRs", "Actual_HRs", "Hit",
    "Insight_Text",   # SHAP-driven explanation (added v2)
]

PERF_COLUMNS = [
    "Date", "Total_Predictions", "High_Conf_Count",
    "Daily_Accuracy", "Cumulative_Accuracy",
    "High_Conf_Hit_Rate", "ROI_Flat_Bet",
]

PLAYER_COLUMNS = [
    "Player", "Team",
    "Season_Predictions", "Season_HRs_Actual", "Season_Hit_Rate",
    "Avg_HR_Prob", "Calibration_Bias",
    "Last_30_Predictions", "Last_30_HRs", "Last_30_Hit_Rate",
    "Last_Updated",
]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HIGH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
MED_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
LOW_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HIT_FILL = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
MISS_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")


# ---------------------------------------------------------------------------
# Workbook initialisation
# ---------------------------------------------------------------------------

def _init_workbook() -> Workbook:
    """Create a new workbook with the four required sheets."""
    wb = Workbook()
    # Default sheet → Predictions
    ws_pred = wb.active
    ws_pred.title = "Predictions"
    _write_header(ws_pred, PRED_COLUMNS)

    ws_perf = wb.create_sheet("Model_Performance")
    _write_header(ws_perf, PERF_COLUMNS)

    ws_feat = wb.create_sheet("Feature_Importance")
    _write_header(ws_feat, ["Date", "Feature", "Importance_Score"])

    ws_player = wb.create_sheet("Player_Stats")
    _write_header(ws_player, PLAYER_COLUMNS)

    wb.save(EXCEL_PATH)
    logger.info("Created %s", EXCEL_PATH)
    return wb


def _write_header(ws, columns: list):
    ws.append(columns)
    for col_idx, _ in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def _get_workbook() -> Workbook:
    if EXCEL_PATH.exists():
        wb = load_workbook(EXCEL_PATH)
        # Migrate older workbooks that predate the Player_Stats sheet
        if "Player_Stats" not in wb.sheetnames:
            ws_player = wb.create_sheet("Player_Stats")
            _write_header(ws_player, PLAYER_COLUMNS)
            wb.save(EXCEL_PATH)
            logger.info("Migrated workbook: added Player_Stats sheet")
        return wb
    return _init_workbook()


# ---------------------------------------------------------------------------
# Save predictions
# ---------------------------------------------------------------------------

def save_predictions(predictions_df: pd.DataFrame):
    """
    Append today's predictions to the Predictions sheet.
    Skips rows that already exist for the same Date+Player+Pitcher combo.
    Applies conditional formatting by confidence tier.
    """
    wb = _get_workbook()
    ws = wb["Predictions"]

    # Build set of already-saved keys to prevent duplicates
    existing_keys: set = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1] and row[4]:  # Date, Player, Pitcher
            existing_keys.add((str(row[0]), str(row[1]).lower(), str(row[4]).lower()))

    new_count = 0
    for _, row in predictions_df.iterrows():
        key = (
            str(row.get("date", date.today().isoformat())),
            str(row.get("player", "")).lower(),
            str(row.get("pitcher", "")).lower(),
        )
        if key in existing_keys:
            continue
        existing_keys.add(key)

        new_row = [
            row.get("date", date.today().isoformat()),
            row.get("player", ""),
            row.get("team", ""),
            row.get("opponent", ""),
            row.get("pitcher", ""),
            round(float(row.get("hr_probability", 0)), 4),
            row.get("confidence", "Low"),
            row.get("park_factor", 100),
            row.get("temp_f", 72),
            row.get("wind_speed_mph", 0),
            row.get("wind_direction", "calm"),
            bool(row.get("is_indoor", False)),
            bool(row.get("home_game", False)),
            round(float(row.get("hr_probability", 0)) * 0.6, 3),  # simple regression proxy
            "",    # Actual_HRs — filled after game
            "",    # Hit — filled after game
            row.get("insight_text", ""),  # SHAP explanation
        ]
        ws.append(new_row)

        # Color-code by confidence
        row_idx = ws.max_row
        fill = (HIGH_FILL if row.get("confidence") == "High"
                else MED_FILL if row.get("confidence") == "Medium"
                else LOW_FILL)
        for col in range(1, len(PRED_COLUMNS) + 1):
            ws.cell(row=row_idx, column=col).fill = fill

    _auto_width(ws)
    wb.save(EXCEL_PATH)
    logger.info("Saved %d predictions → %s", len(predictions_df), EXCEL_PATH)


# ---------------------------------------------------------------------------
# Update actual results
# ---------------------------------------------------------------------------

def update_results(results: list):
    """
    Update actual HR results after games complete.

    Matches results to existing prediction rows by (player_lower, date).
    For HR hitters who have no prediction row (e.g. rookies/prospects missing
    from FanGraphs history), a stub row is appended so they appear in the
    season HR leaderboard.

    Args:
        results: list of dicts {player, date, actual_hrs, team}
    """
    wb = _get_workbook()
    ws = wb["Predictions"]

    result_map = {
        (r["player"].lower(), r["date"]): r
        for r in results
    }

    # Track which (player, date) keys already have prediction rows
    matched: set = set()

    for row in ws.iter_rows(min_row=2):
        player = str(row[1].value or "").lower()
        game_date = str(row[0].value or "")
        key = (player, game_date)

        if key in result_map:
            res = result_map[key]
            actual = res["actual_hrs"]
            hr_prob = float(row[5].value or 0)
            confidence = str(row[6].value or "Low")

            row[14].value = actual  # Actual_HRs
            hit = int(actual) > 0
            row[15].value = int(hit)
            matched.add(key)

            # Color hit/miss for high confidence
            if confidence in ("High", "Medium"):
                fill = HIT_FILL if hit else MISS_FILL
                for cell in row:
                    cell.fill = fill

    # Append stub rows for HR hitters who had no prediction row
    hr_hitters = [r for r in results if r["actual_hrs"] > 0]
    stubs_added = 0
    for r in hr_hitters:
        key = (r["player"].lower(), r["date"])
        if key in matched:
            continue  # Already updated above
        # Only add stub once even if called multiple times
        stub = [
            r["date"],
            r["player"],
            r.get("team", ""),
            "",    # Opponent
            "",    # Pitcher
            0.0,   # HR_Probability (not predicted)
            "N/A", # Confidence
            100,   # Park_Factor (neutral)
            72,    # Temp_F (neutral)
            0,     # Wind_Speed_MPH
            "calm",
            False, # Is_Indoor
            False, # Home_Game
            0.0,   # Predicted_HRs
            r["actual_hrs"],  # Actual_HRs
            1,     # Hit
            "",    # Insight_Text (not predicted)
        ]
        ws.append(stub)
        matched.add(key)
        stubs_added += 1

    if stubs_added:
        logger.info("Added %d stub rows for unmatched HR hitters", stubs_added)

    wb.save(EXCEL_PATH)
    logger.info("Updated results for %d players", len(results))

    # Recalculate daily performance and per-player stats
    _recalculate_performance(wb)
    _rebuild_player_stats(wb)
    wb.save(EXCEL_PATH)


# ---------------------------------------------------------------------------
# Performance recalculation
# ---------------------------------------------------------------------------

def _recalculate_performance(wb: Workbook):
    """Recompute Model_Performance sheet from Predictions sheet."""
    ws_pred = wb["Predictions"]
    ws_perf = wb["Model_Performance"]

    # Clear existing data (keep header)
    for row in ws_perf.iter_rows(min_row=2):
        for cell in row:
            cell.value = None

    # Read predictions into DataFrame (handle old files without Insight_Text column)
    data = list(ws_pred.iter_rows(min_row=2, values_only=True))
    if not data:
        return

    row_len = len(data[0]) if data else 0
    cols = PRED_COLUMNS[:row_len] if row_len <= len(PRED_COLUMNS) else PRED_COLUMNS
    df = pd.DataFrame(data, columns=cols)
    for missing in PRED_COLUMNS:
        if missing not in df.columns:
            df[missing] = ""
    df = df[df["Actual_HRs"].notna()].copy()
    df["Actual_HRs"] = pd.to_numeric(df["Actual_HRs"], errors="coerce").fillna(0)
    df["Hit"] = pd.to_numeric(df["Hit"], errors="coerce").fillna(0)

    if df.empty:
        return

    cumulative_hits = 0
    cumulative_total = 0

    for game_date, day_df in df.groupby("Date"):
        total = len(day_df)
        hits = day_df["Hit"].sum()
        high_df = day_df[day_df["Confidence"] == "High"]
        high_count = len(high_df)
        high_hits = high_df["Hit"].sum()

        cumulative_hits += hits
        cumulative_total += total

        daily_acc = hits / total if total > 0 else 0
        cum_acc = cumulative_hits / cumulative_total if cumulative_total > 0 else 0
        high_hit_rate = high_hits / high_count if high_count > 0 else 0
        roi = (high_hits * 0.9 - (high_count - high_hits)) / high_count if high_count > 0 else 0

        ws_perf.append([
            game_date, total, high_count,
            round(daily_acc, 4), round(cum_acc, 4),
            round(high_hit_rate, 4), round(roi, 4),
        ])

    logger.info("Recalculated performance metrics")


# ---------------------------------------------------------------------------
# Player stats rebuild
# ---------------------------------------------------------------------------

def _rebuild_player_stats(wb: Workbook):
    """Rebuild Player_Stats sheet from Predictions sheet.

    Each row = one unique player with:
    - Season aggregates (all resolved predictions)
    - Last-30-day form
    - Calibration bias (model over/underconfidence per player)
    """
    from datetime import date, timedelta

    ws_pred = wb["Predictions"]

    # Ensure sheet exists
    if "Player_Stats" not in wb.sheetnames:
        ws_player = wb.create_sheet("Player_Stats")
        _write_header(ws_player, PLAYER_COLUMNS)
    else:
        ws_player = wb["Player_Stats"]

    # Clear existing data rows (keep header)
    for row in ws_player.iter_rows(min_row=2):
        for cell in row:
            cell.value = None

    # Load predictions into DataFrame
    data = list(ws_pred.iter_rows(min_row=2, values_only=True))
    if not data:
        return

    row_len = len(data[0]) if data else 0
    cols = PRED_COLUMNS[:row_len] if row_len <= len(PRED_COLUMNS) else PRED_COLUMNS
    df = pd.DataFrame(data, columns=cols)
    for missing in PRED_COLUMNS:
        if missing not in df.columns:
            df[missing] = ""

    df["Date"]         = pd.to_datetime(df["Date"], errors="coerce")
    df["Actual_HRs"]   = pd.to_numeric(df["Actual_HRs"], errors="coerce")
    df["Hit"]          = pd.to_numeric(df["Hit"], errors="coerce")
    df["HR_Probability"] = pd.to_numeric(df["HR_Probability"], errors="coerce")

    # Only resolved rows (Actual_HRs filled) and real predictions (not stubs)
    resolved = df[df["Actual_HRs"].notna() & df["Player"].notna()].copy()
    if resolved.empty:
        return

    cutoff_30 = pd.Timestamp(date.today() - timedelta(days=30))
    last_30   = resolved[resolved["Date"] >= cutoff_30]
    today_str = date.today().isoformat()

    # Season aggregates per player
    season_grp = resolved.groupby("Player").agg(
        Team              = ("Team", "last"),
        Season_Predictions= ("Hit", "count"),
        Season_HRs_Actual = ("Actual_HRs", "sum"),
        Season_Hit_Rate   = ("Hit", "mean"),
        Avg_HR_Prob       = ("HR_Probability", "mean"),
    ).reset_index()
    season_grp["Season_HRs_Actual"] = season_grp["Season_HRs_Actual"].fillna(0).astype(int)
    season_grp["Calibration_Bias"]  = (
        season_grp["Avg_HR_Prob"] - season_grp["Season_Hit_Rate"]
    ).round(4)

    # Last-30-day aggregates per player
    last30_grp = last_30.groupby("Player").agg(
        Last_30_Predictions = ("Hit", "count"),
        Last_30_HRs         = ("Actual_HRs", "sum"),
        Last_30_Hit_Rate    = ("Hit", "mean"),
    ).reset_index()
    last30_grp["Last_30_HRs"] = last30_grp["Last_30_HRs"].fillna(0).astype(int)

    merged = season_grp.merge(last30_grp, on="Player", how="left")
    merged["Last_30_Predictions"] = merged["Last_30_Predictions"].fillna(0).astype(int)
    merged["Last_30_HRs"]        = merged["Last_30_HRs"].fillna(0).astype(int)
    merged["Last_30_Hit_Rate"]   = merged["Last_30_Hit_Rate"].fillna(0.0)
    merged["Last_Updated"]       = today_str

    # Sort by season hit rate descending
    merged = merged.sort_values("Season_Hit_Rate", ascending=False)

    # Write rows
    for _, row in merged.iterrows():
        ws_player.append([
            row["Player"],
            row["Team"],
            int(row["Season_Predictions"]),
            int(row["Season_HRs_Actual"]),
            round(float(row["Season_Hit_Rate"]), 4),
            round(float(row["Avg_HR_Prob"]), 4),
            round(float(row["Calibration_Bias"]), 4),
            int(row["Last_30_Predictions"]),
            int(row["Last_30_HRs"]),
            round(float(row["Last_30_Hit_Rate"]), 4),
            row["Last_Updated"],
        ])

    _auto_width(ws_player)
    logger.info("Rebuilt Player_Stats — %d players", len(merged))


# ---------------------------------------------------------------------------
# Feature importance update
# ---------------------------------------------------------------------------

def update_feature_importance(importance_dict: dict):
    """Append today's feature importance to the Feature_Importance sheet."""
    wb = _get_workbook()
    ws = wb["Feature_Importance"]
    today = date.today().isoformat()
    for feature, score in importance_dict.items():
        ws.append([today, feature, round(float(score), 6)])
    wb.save(EXCEL_PATH)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    # Demo: initialize workbook
    _init_workbook()
    print(f"Initialized {EXCEL_PATH}")

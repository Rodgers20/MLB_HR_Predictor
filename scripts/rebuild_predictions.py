"""
Rebuild corrupted prediction rows.

Old predictions (before the dampening/cap fix) stored raw XGBoost probabilities
(70–99%) instead of calibrated per-game probabilities (≤20%).  This script:

1. Removes all prediction rows with HR_Probability > 0.21 (keeping stub N/A rows)
2. Preserves the actual HR results from those dates as standalone stub rows
3. Re-generates predictions for the affected dates with the fixed predictor
4. Merges results back in
5. Recalculates Model_Performance sheet
"""

import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import logging
import warnings
from copy import deepcopy
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

EXCEL_PATH = Path("MLB_HR_Predictions.xlsx")

PRED_COLUMNS = [
    "Date", "Player", "Team", "Opponent", "Pitcher",
    "HR_Probability", "Confidence", "Park_Factor",
    "Temp_F", "Wind_Speed_MPH", "Wind_Direction", "Is_Indoor",
    "Home_Game", "Predicted_HRs", "Actual_HRs", "Hit",
]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HIGH_FILL   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
MED_FILL    = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
LOW_FILL    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HIT_FILL    = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
MISS_FILL   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")


# ---------------------------------------------------------------------------
# Step 1: load workbook, identify bad rows, rescue actual-result stubs
# ---------------------------------------------------------------------------

logger.info("Loading workbook ...")
wb   = load_workbook(EXCEL_PATH)
ws   = wb["Predictions"]

all_rows   = list(ws.iter_rows(min_row=2, values_only=True))
df_all     = pd.DataFrame(all_rows, columns=PRED_COLUMNS)
df_all     = df_all[df_all["Player"].notna()].copy()
df_all["HR_Probability"] = pd.to_numeric(df_all["HR_Probability"], errors="coerce")

bad_mask  = df_all["HR_Probability"] > 0.21
good_mask = ~bad_mask

df_good = df_all[good_mask].copy()
df_bad  = df_all[bad_mask].copy()

logger.info("Good rows: %d  |  Bad rows (old model): %d", len(df_good), len(df_bad))
logger.info("Bad dates: %s", sorted(df_bad["Date"].astype(str).unique()))

# Rescue actual HR results from bad rows that had results filled in
actual_results = {}
for _, row in df_bad.iterrows():
    if row["Actual_HRs"] not in (None, "", "None"):
        try:
            actual = int(float(row["Actual_HRs"]))
        except (ValueError, TypeError):
            continue
        if actual > 0:
            key = (str(row["Player"]).lower(), str(row["Date"]))
            if key not in actual_results:
                actual_results[key] = {
                    "player": row["Player"],
                    "team":   row["Team"],
                    "date":   str(row["Date"]),
                    "actual_hrs": actual,
                }

logger.info("Rescued %d actual-HR results from bad rows", len(actual_results))


# ---------------------------------------------------------------------------
# Step 2: clear Predictions sheet (keep header), write good rows
# ---------------------------------------------------------------------------

for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.value = None

# Rewrite good rows
for _, row in df_good.iterrows():
    ws.append(row.tolist())

logger.info("Rewritten %d good rows to Predictions sheet", len(df_good))


# ---------------------------------------------------------------------------
# Step 3: regenerate predictions for bad dates using the fixed predictor
# ---------------------------------------------------------------------------

bad_dates = sorted(df_bad["Date"].astype(str).str[:10].unique())
logger.info("Regenerating predictions for: %s", bad_dates)

from utils.predictor import predict_today
from tracker.prediction_tracker import save_predictions, update_results, _recalculate_performance

# Save workbook before running predictor (predictor reads xlsx via load_workbook)
wb.save(EXCEL_PATH)

for game_date in bad_dates:
    logger.info("--- Predicting %s ---", game_date)
    try:
        preds = predict_today(game_date)
        if preds.empty:
            logger.warning("No predictions for %s", game_date)
            continue
        save_predictions(preds)
        logger.info("Saved %d predictions for %s", len(preds), game_date)
    except Exception as exc:
        logger.error("Failed to predict %s: %s", game_date, exc)


# ---------------------------------------------------------------------------
# Step 4: re-apply rescued actual results
# ---------------------------------------------------------------------------

if actual_results:
    logger.info("Re-applying %d actual results ...", len(actual_results))
    update_results(list(actual_results.values()))
    logger.info("Results applied")
else:
    # Still recalculate performance from whatever results exist
    wb2 = load_workbook(EXCEL_PATH)
    _recalculate_performance(wb2)
    wb2.save(EXCEL_PATH)

logger.info("Done — workbook rebuilt at %s", EXCEL_PATH)

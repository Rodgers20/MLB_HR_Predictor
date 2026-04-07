"""
Daily prediction pipeline.
1. Fetch today's MLB schedule and lineups via mlb-statsapi
2. Build matchup feature matrix for each batter vs starting pitcher
3. Run XGBoost → output HR probability + projected HRs
4. Flag Best Bets by confidence tier
"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import logging
from datetime import date
from pathlib import Path

import numpy as np
import pandas as pd
import statsapi

from utils.data_fetcher import (
    fetch_fangraphs_batting,
    fetch_fangraphs_pitching,
    load_park_factors,
)
from utils.feature_engineer import (
    MODEL_FEATURES,
    build_3yr_weighted_fg,
    build_3yr_weighted_pitcher,
    build_matchup_features,
)
from utils.model_trainer import load_model
from utils.weather_fetcher import fetch_all_game_weather
from utils.roster_fetcher import get_current_roster_map
from utils.explainer import explain_prediction

logger = logging.getLogger(__name__)

HIGH_CONF = float(os.getenv("HIGH_CONFIDENCE_THRESHOLD", 0.18))
MED_CONF  = float(os.getenv("MEDIUM_CONFIDENCE_THRESHOLD", 0.12))

EXCEL_PATH = Path("MLB_HR_Predictions.xlsx")
# Minimum predictions before bias correction is applied for a player
_MIN_PREDS_FOR_CORRECTION = 10
# Maximum fractional correction per player (caps over/under-adjustment)
_MAX_CORRECTION = 0.20


def _load_player_bias_map() -> dict:
    """Return {player_name_lower: calibration_bias} from Player_Stats sheet.

    calibration_bias = avg_predicted_prob - actual_hit_rate.
    Positive  → model overestimates → we should scale down.
    Negative  → model underestimates → we should scale up.
    Only players with Season_Predictions >= _MIN_PREDS_FOR_CORRECTION are included.
    """
    if not EXCEL_PATH.exists():
        return {}
    try:
        from openpyxl import load_workbook as _lw
        wb  = _lw(EXCEL_PATH, read_only=True, data_only=True)
        if "Player_Stats" not in wb.sheetnames:
            wb.close()
            return {}
        ws  = wb["Player_Stats"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
    except Exception as exc:
        logger.warning("Could not load Player_Stats for bias correction: %s", exc)
        return {}

    bias_map: dict = {}
    for row in rows:
        if len(row) < 7 or row[0] is None:
            continue
        player_name       = str(row[0])
        season_predictions = int(row[2] or 0)
        calibration_bias  = float(row[6] or 0)
        if season_predictions >= _MIN_PREDS_FOR_CORRECTION:
            # Shrink correction toward 0 for small samples
            shrinkage = min(season_predictions / 30.0, 1.0) * 0.5
            bias_map[player_name.lower()] = calibration_bias * shrinkage
    logger.info("Loaded bias corrections for %d players", len(bias_map))
    return bias_map


def _data_year_for_date(game_date: str) -> int:
    """Return the season year for the given game date."""
    from datetime import datetime
    return datetime.strptime(game_date, "%Y-%m-%d").year


# ---------------------------------------------------------------------------
# Team name → abbreviation
# ---------------------------------------------------------------------------

TEAM_NAME_TO_ABBR = {
    "Arizona Diamondbacks": "ARI", "Atlanta Braves": "ATL",
    "Baltimore Orioles": "BAL", "Boston Red Sox": "BOS",
    "Chicago Cubs": "CHC", "Chicago White Sox": "CHW",
    "Cincinnati Reds": "CIN", "Cleveland Guardians": "CLE",
    "Colorado Rockies": "COL", "Detroit Tigers": "DET",
    "Houston Astros": "HOU", "Kansas City Royals": "KC",
    "Los Angeles Angels": "LAA", "Los Angeles Dodgers": "LAD",
    "Miami Marlins": "MIA", "Milwaukee Brewers": "MIL",
    "Minnesota Twins": "MIN", "New York Mets": "NYM",
    "New York Yankees": "NYY", "Oakland Athletics": "OAK",
    "Philadelphia Phillies": "PHI", "Pittsburgh Pirates": "PIT",
    "San Diego Padres": "SD", "San Francisco Giants": "SF",
    "Seattle Mariners": "SEA", "St. Louis Cardinals": "STL",
    "Tampa Bay Rays": "TB", "Texas Rangers": "TEX",
    "Toronto Blue Jays": "TOR", "Washington Nationals": "WSH",
    "Athletics": "OAK",
}


def _abbr(full_name: str) -> str:
    for key, abbr in TEAM_NAME_TO_ABBR.items():
        if key in full_name:
            return abbr
    return full_name[:3].upper()


# ---------------------------------------------------------------------------
# Schedule + lineup helpers
# ---------------------------------------------------------------------------

def fetch_todays_games(game_date: str = None) -> list:
    if game_date is None:
        game_date = date.today().strftime("%Y-%m-%d")
    try:
        schedule = statsapi.schedule(date=game_date)
    except Exception as exc:
        logger.error("Failed to fetch schedule: %s", exc)
        return []

    games = []
    for g in schedule:
        games.append({
            "game_id": g.get("game_id"),
            "home_team": _abbr(g.get("home_name", "")),
            "away_team": _abbr(g.get("away_name", "")),
            "game_date": game_date,
            "game_hour_utc": 19,
            "probable_home_pitcher": g.get("home_probable_pitcher", "TBD"),
            "probable_away_pitcher": g.get("away_probable_pitcher", "TBD"),
        })
    logger.info("Found %d games on %s", len(games), game_date)
    return games


def fetch_lineups(game_id: int) -> dict:
    try:
        game = statsapi.get("game", {"gamePk": game_id})
        lineups = {"home": [], "away": []}
        box = game.get("liveData", {}).get("boxscore", {}).get("teams", {})
        for side in ("home", "away"):
            players = box.get(side, {}).get("battingOrder", [])
            for pid in players:
                player_info = box[side]["players"].get(f"ID{pid}", {})
                name = player_info.get("person", {}).get("fullName", "")
                if name:
                    lineups[side].append(name)
        return lineups
    except Exception as exc:
        logger.warning("Could not fetch lineup for game %s: %s", game_id, exc)
        return {"home": [], "away": []}


# ---------------------------------------------------------------------------
# Load player data (current year + 2 prior for 3yr weighted)
# ---------------------------------------------------------------------------

def _load_player_data(current_year: int) -> tuple:
    fg_bat_frames, fg_pit_frames = [], []
    for yr in range(current_year - 2, current_year + 1):
        try:
            fg_bat_frames.append(fetch_fangraphs_batting(yr))
            fg_pit_frames.append(fetch_fangraphs_pitching(yr))
        except Exception as exc:
            logger.warning("Could not load year %d: %s", yr, exc)

    fg_batting  = pd.concat([d for d in fg_bat_frames  if d is not None], ignore_index=True)
    fg_pitching = pd.concat([d for d in fg_pit_frames  if d is not None], ignore_index=True)
    return fg_batting, fg_pitching


# ---------------------------------------------------------------------------
# Main prediction function
# ---------------------------------------------------------------------------

def predict_today(game_date: str = None) -> pd.DataFrame:
    if game_date is None:
        game_date = date.today().strftime("%Y-%m-%d")

    logger.info("Running predictions for %s ...", game_date)

    model, scaler   = load_model()
    park_factors    = load_park_factors()
    park_factor_map = dict(zip(park_factors["team"], park_factors["hr_park_factor"]))
    player_bias_map = _load_player_bias_map()

    current_year    = _data_year_for_date(game_date)
    logger.info("Using player data year: %d", current_year)
    fg_batting, fg_pitching = _load_player_data(current_year)

    # Live 2026 roster map: {player_name_lower: current_team_abbr}
    # Used to (1) build correct batter pools when no lineup is confirmed,
    # and (2) override stale FanGraphs team assignments in output.
    game_season = int(game_date[:4])
    roster_map  = get_current_roster_map(game_season)
    logger.info("Loaded roster map: %d players for %d season", len(roster_map), game_season)

    # Current-year slices (FanGraphs uses capital "Season")
    year_col   = "Season" if "Season" in fg_batting.columns else "season"
    fg_bat_cur = fg_batting[fg_batting[year_col] == current_year].copy()
    fg_pit_cur = fg_pitching[fg_pitching[year_col] == current_year].copy()

    # Early-season fallback: if FanGraphs hasn't published enough current-year
    # data yet (< 100 qualified batters), use prior season for model features so
    # predictions remain meaningful.  Display stats in the dashboard still try
    # current year first (handled separately in _lookup_batting_stats).
    MIN_BATTERS = 100
    if len(fg_bat_cur) < MIN_BATTERS:
        logger.info(
            "Current year %d has only %d batters — falling back to %d for model features",
            current_year, len(fg_bat_cur), current_year - 1,
        )
        current_year    = current_year - 1
        fg_batting, fg_pitching = _load_player_data(current_year)
        year_col        = "Season" if "Season" in fg_batting.columns else "season"
        fg_bat_cur      = fg_batting[fg_batting[year_col] == current_year].copy()
        fg_pit_cur      = fg_pitching[fg_pitching[year_col] == current_year].copy()

    fg_bat_cur["hr_rate"] = fg_bat_cur["HR"] / fg_bat_cur["PA"].replace(0, np.nan)

    # 3-year weighted batter features — use last *completed* season so partial
    # 2026 data (< 40 PA) doesn't drag down elite hitters like Cal Raleigh.
    # Bayesian smoothing in build_matchup_features() handles the current year.
    batter_weighted  = build_3yr_weighted_fg(fg_batting, current_year - 1)
    pitcher_weighted = build_3yr_weighted_pitcher(fg_pitching, current_year - 1)

    games = fetch_todays_games(game_date)
    if not games:
        logger.warning("No games found for %s", game_date)
        return pd.DataFrame()

    weather_map  = fetch_all_game_weather(games)
    predictions  = []

    for game in games:
        home        = game["home_team"]
        away        = game["away_team"]
        weather     = weather_map.get(home, {"temp_f": 72.0, "wind_speed_mph": 0.0,
                                              "wind_direction": "calm", "is_indoor": False})
        park_factor = park_factor_map.get(home, 100.0)
        lineups     = fetch_lineups(game["game_id"]) if game.get("game_id") else {}

        sides = [
            (home, away, game["probable_away_pitcher"], lineups.get("home", [])),
            (away, home, game["probable_home_pitcher"], lineups.get("away", [])),
        ]

        for batting_team, fielding_team, pitcher_name, batters in sides:
            # Pitcher row (current season FanGraphs)
            pit_match = fg_pit_cur[
                fg_pit_cur["Name"].str.contains(pitcher_name, case=False, na=False)
            ] if pitcher_name and pitcher_name != "TBD" else pd.DataFrame()
            pitcher_series = pit_match.iloc[0] if not pit_match.empty else pd.Series()

            # Batter pool: confirmed lineup > live roster > FanGraphs team fallback
            if batters:
                # Lineup confirmed — match by name
                batter_pool = fg_bat_cur[
                    fg_bat_cur["Name"].apply(
                        lambda n: any(b.lower() in str(n).lower() for b in batters)
                    )
                ]
            elif roster_map:
                # Use live 2026 roster to find players currently on this team,
                # then match those names against FanGraphs historical stats.
                current_team_players = {
                    name for name, abbr in roster_map.items()
                    if abbr == batting_team
                }
                batter_pool = fg_bat_cur[
                    fg_bat_cur["Name"].apply(
                        lambda n: str(n).lower() in current_team_players
                        or any(
                            cp in str(n).lower() or str(n).lower() in cp
                            for cp in current_team_players
                            if len(cp) > 5  # skip very short names
                        )
                    )
                ]
                if batter_pool.empty:
                    # Final fallback: old FanGraphs team column
                    batter_pool = fg_bat_cur[fg_bat_cur["Team"] == batting_team]
            else:
                batter_pool = fg_bat_cur[fg_bat_cur["Team"] == batting_team]

            for _, batter_series in batter_pool.iterrows():
                batter_id = batter_series.get("IDfg")

                # Weighted batter features
                bw_row = batter_weighted[batter_weighted["IDfg"] == batter_id]
                batter_w = bw_row.iloc[0] if not bw_row.empty else pd.Series()

                # Weighted pitcher features
                pitcher_fg_id = pitcher_series.get("IDfg") if not pitcher_series.empty else None
                pw_row = pitcher_weighted[pitcher_weighted["IDfg"] == pitcher_fg_id] \
                         if pitcher_fg_id is not None else pd.DataFrame()
                # Merge current + weighted pitcher into one Series
                combined_pitcher = pd.concat([
                    pitcher_series,
                    pw_row.iloc[0] if not pw_row.empty else pd.Series(),
                ])

                features = build_matchup_features(
                    batter_row     = batter_series,
                    batter_weighted= batter_w,
                    pitcher_row    = combined_pitcher,
                    park_factor    = park_factor,
                    temp_f         = weather["temp_f"],
                    wind_speed_mph = weather["wind_speed_mph"],
                    wind_direction = weather["wind_direction"],
                    is_day_game    = False,
                )

                feat_df = pd.DataFrame([features])[MODEL_FEATURES]
                feat_df = feat_df.fillna(0)
                feat_scaled = scaler.transform(feat_df)

                # ── Model output ───────────────────────────────────────────
                # Model: XGBoost trained on season-level "elite power hitter"
                # classification (HR/PA >= 3.5%).  Calibrated output range 0–1.
                # power_score of 0.9 → very strong Statcast profile
                # power_score of 0.5 → league-average profile
                # power_score of 0.1 → weak contact profile
                power_score = float(model.predict_proba(feat_scaled)[0][1])

                # ── Bayesian shrinkage of observed HR rate ─────────────────
                # Prevent small-sample flukes (e.g. 4 HR in 47 PA) from
                # outranking proven sluggers.  Blend toward league avg at low PA.
                LEAGUE_HR_RATE = 0.025   # MLB average HR/PA
                MIN_PA_FULL    = 150     # PA needed to fully trust observed hr_rate
                PA_PER_GAME    = 3.5     # avg plate appearances per game

                pa      = float(batter_series.get("PA", 0) or 0)
                hr_rate = float(batter_series.get("hr_rate", 0) or 0)

                if pa >= MIN_PA_FULL:
                    trusted_hr_rate = hr_rate if hr_rate > 0 else LEAGUE_HR_RATE
                elif pa > 0:
                    w = pa / MIN_PA_FULL              # 0 → 1 as pa → 150
                    trusted_hr_rate = hr_rate * w + LEAGUE_HR_RATE * (1 - w)
                else:
                    trusted_hr_rate = LEAGUE_HR_RATE

                # ── Per-game probability components ────────────────────────
                # Component 1: Observed HR rate scaled to per-game
                rate_component = trusted_hr_rate * PA_PER_GAME

                # Component 2: Model Statcast profile → per-game probability
                # Exponential mapping: 0.5 → 9%, 0.9 → 22%, 0.1 → 3.5%
                # This creates real separation between elite and average hitters.
                model_component = 0.09 * float(np.exp(2.5 * (power_score - 0.5)))

                # Blend: high PA → trust observed rate more; low PA → trust model
                # pa_blend reaches max 0.55 when pa ≥ MIN_PA_FULL
                pa_blend = min(pa / MIN_PA_FULL, 1.0) * 0.55
                base_prob = rate_component * pa_blend + model_component * (1.0 - pa_blend)

                # ── Context multipliers ────────────────────────────────────
                # Weather features are constant in training (temp=72, wind=0)
                # → model can't learn them → apply manually here
                _temp_f   = float(weather.get("temp_f", 72) or 72)
                _wind_dir = str(weather.get("wind_direction", "calm") or "calm").lower()

                temp_adj = 1.0 + max((_temp_f - 72.0), -15.0) / 10.0 * 0.012
                wind_adj = {"out": 1.12, "in": 0.88, "cross": 0.97, "calm": 1.0}.get(_wind_dir, 1.0)
                park_adj = (park_factor / 100.0) ** 0.5

                hr_prob = base_prob * park_adj * temp_adj * wind_adj

                # ── Player-level bias correction ───────────────────────────
                # If we've tracked this player's results before, nudge the
                # probability toward reality.  bias > 0 means we've been
                # overconfident → scale down; bias < 0 → scale up.
                _pname_lower = batter_series.get("Name", "").lower()
                if _pname_lower in player_bias_map:
                    bias         = player_bias_map[_pname_lower]
                    correction   = float(np.clip(-bias / max(hr_prob, 0.01), -_MAX_CORRECTION, _MAX_CORRECTION))
                    hr_prob     *= (1.0 + correction)

                hr_prob = float(np.clip(hr_prob, 0.005, 0.35))  # cap raised to 35%

                confidence = (
                    "High"   if hr_prob >= HIGH_CONF else
                    "Medium" if hr_prob >= MED_CONF  else
                    "Low"
                )

                # ── SHAP insight text ──────────────────────────────────────
                # Explain WHY the model chose this player in plain English
                insight_text, _ = explain_prediction(
                    model=model,
                    feat_scaled_row=feat_scaled,
                    feature_names=MODEL_FEATURES,
                    feat_values_dict=features,
                )

                player_name = batter_series.get("Name", "Unknown")
                # Use live roster team; fall back to game schedule team
                live_team = (
                    roster_map.get(player_name.lower())
                    or batting_team
                )

                predictions.append({
                    "date":           game_date,
                    "player":         player_name,
                    "team":           live_team,
                    "opponent":       fielding_team,
                    "pitcher":        pitcher_name,
                    "hr_probability": round(hr_prob, 4),
                    "confidence":     confidence,
                    "park_factor":    park_factor,
                    "temp_f":         weather["temp_f"],
                    "wind_speed_mph": weather["wind_speed_mph"],
                    "wind_direction": weather["wind_direction"],
                    "is_indoor":      weather.get("is_indoor", False),
                    "home_game":      batting_team == home,
                    "insight_text":   insight_text,
                })

    df = pd.DataFrame(predictions)
    if not df.empty:
        df = df.sort_values("hr_probability", ascending=False).reset_index(drop=True)
        logger.info("Generated %d predictions (%d High confidence)",
                    len(df), (df["confidence"] == "High").sum())
    return df


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import argparse
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

    parser = argparse.ArgumentParser()
    parser.add_argument("--date", type=str, default=None,
                        help="Game date YYYY-MM-DD (default: today)")
    parser.add_argument("--save", action="store_true",
                        help="Save predictions to Excel tracker")
    args = parser.parse_args()

    preds = predict_today(args.date)
    if not preds.empty:
        print(preds[["player", "team", "pitcher", "hr_probability", "confidence",
                      "park_factor", "wind_direction"]].to_string(index=False))
        if args.save:
            from tracker.prediction_tracker import save_predictions
            save_predictions(preds)
            print("Saved to MLB_HR_Predictions.xlsx")
    else:
        print("No predictions generated.")
